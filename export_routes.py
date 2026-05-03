"""
export_routes.py — Echo Audit V2 Phase 6 data export + restore.

Three routes:
    GET  /api/export/interactions  — Excel (.xlsx) download
    GET  /api/export/backup        — Full-tenant JSON backup
    POST /api/export/restore       — Restore a JSON backup (all-or-nothing)

Every route is company-scoped; the super-admin platform uses separate
platform-level routes. Credentials, password hashes, and encryption keys
are NEVER included in export output (see `_SANITIZED_USER_COLUMNS` and the
VoIP backup block that drops `voip_config_credentials`).
"""

import base64
import io
import json
import logging
from datetime import date, datetime

from flask import Blueprint, Response, jsonify, request, send_file, stream_with_context
from flask_login import current_user, login_required

from audit_log import ACTION_EXPORTED, ACTION_UPDATED, ENTITY_COMPANY, write_audit_log
from auth import role_required
from db import IS_POSTGRES, get_conn, q
from helpers import get_effective_company_id

logger = logging.getLogger(__name__)

export_bp = Blueprint("export", __name__, url_prefix="/api/export")


def _err(msg, code):
    return jsonify({"error": msg}), code


def _require_company():
    cid = get_effective_company_id()
    if cid is None:
        return None, _err(
            "No company context. Super admins must select an organization first.",
            400,
        )
    return cid, None


def _row_to_dict(row):
    if row is None:
        return None
    try:
        return dict(row)
    except Exception:
        return {k: row[k] for k in row.keys()}


def _rows(cur):
    return [_row_to_dict(r) for r in cur.fetchall()]


def _json_default(value):
    """Serialize date/datetime/memoryview for JSON output."""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, memoryview):
        # Avoid dumping raw bytes in JSON backups — BYTEA blobs are excluded
        # intentionally at the SELECT level, but this is a safety net.
        return None
    if isinstance(value, bytes):
        return None
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")


# ═══════════════════════════════════════════════════════════════
# GET /api/export/interactions  — Excel (.xlsx)
# ═══════════════════════════════════════════════════════════════


@export_bp.route("/interactions", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def export_interactions():
    """NDJSON-streaming Excel export for a user-selected set of interactions.

    Body: {"interaction_ids": [int, ...]}  (1-50, all in current company).

    Response: application/x-ndjson, one JSON object per line:
        {"type":"progress","current":N,"total":M,"label":"..."}
        {"type":"done","filename":"...","xlsx_base64":"..."}
        {"type":"error","message":"..."}

    AI Call Summary is freshly generated per row by Haiku — no caching.
    Streaming output keeps the gunicorn worker alive past the 120s timeout
    without a Procfile change (timeout is no-progress, not total-request).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _err("openpyxl is not installed on the server", 500)

    company_id, err = _require_company()
    if err: return err

    body = request.get_json(silent=True) or {}
    raw_ids = body.get("interaction_ids") or []
    try:
        interaction_ids = [int(x) for x in raw_ids]
    except (TypeError, ValueError):
        return _err("interaction_ids must be a list of integers", 400)
    if not interaction_ids:
        return _err("Select at least one interaction to export.", 400)
    if len(interaction_ids) > 50:
        return _err("Maximum 50 interactions per export.", 400)

    # Pre-validate company scope before opening the stream so auth-style
    # errors come back as proper HTTP errors instead of mid-stream JSON.
    placeholders = ",".join(["?"] * len(interaction_ids))
    conn = get_conn()
    try:
        cur = conn.execute(
            q(f"""SELECT i.interaction_id
                    FROM interactions i
                    JOIN projects p ON p.project_id = i.project_id
                   WHERE p.company_id = ?
                     AND i.interaction_deleted_at IS NULL
                     AND i.interaction_id IN ({placeholders})"""),
            [company_id] + interaction_ids,
        )
        valid_ids = {r["interaction_id"] for r in cur.fetchall()}
    finally:
        conn.close()
    missing = [iid for iid in interaction_ids if iid not in valid_ids]
    if missing:
        return _err(
            f"{len(missing)} interaction(s) not found or not in your company.", 404,
        )

    actor_user_id = current_user.user_id

    def _generate():
        from grader import summarize_call_for_export, summarize_no_answer_for_export

        try:
            conn = get_conn()
            try:
                # Universal call_when via COALESCE chain spanning all 3 sources:
                #   AI Shop  → scheduled_calls.sc_requested_at
                #   Live VoIP → voip_call_queue.voip_queue_created_at
                #   Live Record (browser) → interaction_call_start_time
                #   Upload   → interaction_uploaded_at
                #   Fallback → interaction_submitted_at
                cur = conn.execute(
                    q(f"""SELECT
                            i.interaction_id,
                            i.status_id,
                            i.interaction_overall_score,
                            i.interaction_transcript,
                            i.interaction_responder_name,
                            loc.location_name,
                            COALESCE(
                                r.respondent_name,
                                NULLIF(TRIM(respondent.user_first_name || ' '
                                            || respondent.user_last_name), ''),
                                NULLIF(TRIM(i.interaction_responder_name), ''),
                                'Name Not Detected'
                            )                                       AS respondent_display,
                            COALESCE(
                                sc.sc_requested_at,
                                vcq.voip_queue_created_at,
                                i.interaction_call_start_time,
                                i.interaction_uploaded_at,
                                i.interaction_submitted_at
                            )                                       AS call_when
                          FROM interactions i
                          JOIN projects p              ON p.project_id     = i.project_id
                          LEFT JOIN locations  loc     ON loc.location_id  = i.interaction_location_id
                          LEFT JOIN users respondent   ON respondent.user_id = i.respondent_user_id
                          LEFT JOIN respondents r      ON r.respondent_id  = i.respondent_id
                          LEFT JOIN voip_call_queue vcq ON vcq.voip_queue_interaction_id = i.interaction_id
                          LEFT JOIN scheduled_calls sc  ON sc.sc_conversation_id        = vcq.voip_queue_call_id
                         WHERE p.company_id = ?
                           AND i.interaction_deleted_at IS NULL
                           AND i.interaction_id IN ({placeholders})
                         ORDER BY call_when DESC NULLS LAST, i.interaction_id DESC"""),
                    [company_id] + interaction_ids,
                )
                rows = _rows(cur)

                # Per-interaction rubric scores keyed by snapshot name + union
                # of names for dynamic columns (matches existing pattern).
                scores_by_iid = {r["interaction_id"]: {} for r in rows}
                rubric_columns: list[str] = []
                seen = set()
                cur = conn.execute(
                    q(f"""SELECT interaction_id, irs_snapshot_name, irs_score_value
                            FROM interaction_rubric_scores
                           WHERE interaction_id IN ({placeholders})
                           ORDER BY interaction_rubric_score_id ASC"""),
                    interaction_ids,
                )
                for sr in cur.fetchall():
                    d = _row_to_dict(sr)
                    scores_by_iid.setdefault(d["interaction_id"], {})[d["irs_snapshot_name"]] = \
                        d["irs_score_value"]
                    if d["irs_snapshot_name"] not in seen:
                        seen.add(d["irs_snapshot_name"])
                        rubric_columns.append(d["irs_snapshot_name"])
            finally:
                conn.close()

            wb = Workbook()
            ws = wb.active
            ws.title = "Past Grades"
            header = ["Location", "Date called", "Time called", "Respondent",
                      "Total Score"] + rubric_columns + ["Call Summary"]
            ws.append(header)

            # Center-align score columns on header + every data row. 1-based:
            # Total Score is column 5; rubric columns run 6..(5+N).
            CENTER = Alignment(horizontal="center", vertical="center")
            WRAP   = Alignment(wrap_text=True, vertical="top")
            score_cols = [5] + list(range(6, 6 + len(rubric_columns)))
            for col in score_cols:
                ws.cell(row=1, column=col).alignment = CENTER
            # Bold + freeze header row.
            BOLD = Font(bold=True)
            for col in range(1, ws.max_column + 1):
                ws.cell(row=1, column=col).font = BOLD
            ws.freeze_panes = "A2"

            # Per-row height estimator for wrapped Call Summary cells.
            # Call Summary is capped at width=80; effective wrap ≈ 75 chars
            # per line at default Calibri 11pt; line height ≈ 15pt.
            def _row_height_for_summary(summary, cpl=75, line_h=15, min_h=15):
                if not isinstance(summary, str) or not summary:
                    return min_h
                lines = max(1, (len(summary) + cpl - 1) // cpl)
                return max(min_h, lines * line_h)

            total = len(rows)
            for idx, r in enumerate(rows, start=1):
                yield json.dumps({
                    "type":    "progress",
                    "current": idx,
                    "total":   total,
                    "label":   f"Generating summary {idx}/{total} — {r.get('location_name') or 'Unknown'}",
                }) + "\n"

                cw = r.get("call_when")
                if hasattr(cw, "strftime"):
                    date_str = cw.date().isoformat() if hasattr(cw, "date") else cw.isoformat()[:10]
                    time_str = cw.strftime("%-I:%M %p")
                elif isinstance(cw, str):
                    date_str = cw[:10]
                    time_str = ""
                    if len(cw) >= 16:
                        try:
                            time_str = datetime.fromisoformat(cw[:19]).strftime("%-I:%M %p")
                        except (ValueError, TypeError):
                            time_str = cw[11:16]   # last-resort raw 24-hour
                else:
                    date_str = ""
                    time_str = ""

                row_scores = scores_by_iid.get(r["interaction_id"], {})
                # No-answer rows get a short canonical label instead of a
                # narrative summary — see grader.summarize_no_answer_for_export
                # for the label set + classifier reuse.
                if r.get("status_id") == 44:
                    summary = summarize_no_answer_for_export(
                        r.get("interaction_transcript") or ""
                    )
                else:
                    summary = summarize_call_for_export(
                        transcript=r.get("interaction_transcript") or "",
                        scores_per_criterion=row_scores,
                        location_name=r.get("location_name"),
                        respondent_name=r.get("respondent_display"),
                    )

                total_score = r.get("interaction_overall_score")
                if r.get("status_id") == 44:
                    total_value = "No answer"
                else:
                    total_value = float(total_score) if total_score is not None else ""
                row = [
                    r.get("location_name") or "",
                    date_str,
                    time_str,
                    r.get("respondent_display") or "",
                    total_value,
                ]
                for col in rubric_columns:
                    v = row_scores.get(col)
                    if v is None:
                        row.append(None)
                    else:
                        try:
                            row.append(float(v))
                        except (TypeError, ValueError):
                            # Yes/No/Pending stays as string.
                            row.append(str(v))
                row.append(summary)
                ws.append(row)
                for col in score_cols:
                    ws.cell(row=ws.max_row, column=col).alignment = CENTER
                # Wrap-text + top-align on the Call Summary cell so capped
                # columns don't clip long single-sentence summaries.
                ws.cell(row=ws.max_row, column=ws.max_column).alignment = WRAP
                # Bold Total Score per data row (column 5) so it pops as the
                # at-a-glance focus column. Header is already bold from G2.4.
                ws.cell(row=ws.max_row, column=5).font = BOLD
                # Auto-grow row height to fit the wrapped Call Summary so
                # users see the whole sentence without expanding the row.
                ws.row_dimensions[ws.max_row].height = _row_height_for_summary(summary)

            # Auto-fit column widths from longest visible value (header +
            # every data row). Call Summary is capped at 80 — wrap-text on
            # its data cells handles the overflow into multi-line cells.
            CS_COL = ws.max_column
            for col_idx in range(1, ws.max_column + 1):
                max_len = max(
                    (len(str(ws.cell(row=r, column=col_idx).value))
                     for r in range(1, ws.max_row + 1)
                     if ws.cell(row=r, column=col_idx).value is not None),
                    default=10,
                )
                width = min(max_len + 2, 80) if col_idx == CS_COL else max_len + 2
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Conditional formatting on score columns (data range only).
            # ISNUMBER guard skips empty / Yes-No string cells so no fill
            # appears on no-answer rows or non-numeric criteria.
            GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            if ws.max_row >= 2:
                for col_idx in score_cols:
                    cl  = get_column_letter(col_idx)
                    rng = f"{cl}2:{cl}{ws.max_row}"
                    top = f"{cl}2"
                    ws.conditional_formatting.add(rng, FormulaRule(
                        formula=[f"AND(ISNUMBER({top}),{top}<5)"], fill=RED))
                    ws.conditional_formatting.add(rng, FormulaRule(
                        formula=[f"AND(ISNUMBER({top}),{top}>=5,{top}<7)"], fill=YELLOW))
                    ws.conditional_formatting.add(rng, FormulaRule(
                        formula=[f"AND(ISNUMBER({top}),{top}>=7)"], fill=GREEN))

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            xlsx_b64 = base64.b64encode(buf.read()).decode("ascii")

            try:
                write_audit_log(
                    actor_user_id, ACTION_EXPORTED, ENTITY_COMPANY, company_id,
                    metadata={"action":         "export_interactions_v2",
                              "row_count":      total,
                              "interaction_ids": interaction_ids},
                )
            except Exception:
                logger.exception("export_interactions audit write failed")

            yield json.dumps({
                "type":        "done",
                "filename":    f"echoaudit_export_{date.today().isoformat()}.xlsx",
                "xlsx_base64": xlsx_b64,
            }) + "\n"
        except Exception as exc:
            logger.exception("export_interactions stream failed")
            yield json.dumps({"type": "error", "message": str(exc)}) + "\n"

    return Response(
        stream_with_context(_generate()),
        mimetype="application/x-ndjson",
        headers={
            "Cache-Control":     "no-cache, no-transform",
            "X-Accel-Buffering": "no",  # disable proxy buffering on Railway edge
        },
    )


# ═══════════════════════════════════════════════════════════════
# GET /api/export/backup  — JSON backup
# ═══════════════════════════════════════════════════════════════
#
# Tables included (strictly company-scoped):
#   companies, company_settings, company_labels,
#   locations, departments, phone_routing, rubric_groups, rubric_items,
#   projects, interactions, interaction_rubric_scores,
#   clarifying_questions, performance_reports,
#   voip_call_queue, users
#
# Explicitly excluded:
#   user_password_hash                    — credential material
#   voip_config_credentials / voip_config — stays in DB; restoring on a new
#                                            host needs the same encryption
#                                            key anyway
#   interaction_audio_data / voip_queue_recording_data — binary blobs
#   api_keys / api_usage / api_call_log   — operational, not business data
#   audit_log                             — append-only system record


_SANITIZED_USER_COLUMNS = """
    u.user_id, u.user_role_id, u.department_id, u.user_email,
    u.user_first_name, u.user_last_name, u.status_id,
    u.user_must_change_password, u.user_created_at, u.user_updated_at
"""


@export_bp.route("/backup", methods=["GET"])
@login_required
@role_required("admin", "super_admin")
def export_backup():
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        # 1. companies
        cur = conn.execute(
            q("SELECT * FROM companies WHERE company_id = ?"),
            (company_id,),
        )
        companies = _rows(cur)
        # 2. company_settings
        cur = conn.execute(
            q("SELECT * FROM company_settings WHERE company_id = ?"),
            (company_id,),
        )
        settings = _rows(cur)
        # 3. company_labels
        cur = conn.execute(
            q("SELECT * FROM company_labels WHERE company_id = ?"),
            (company_id,),
        )
        labels = _rows(cur)
        # 4. locations
        cur = conn.execute(
            q("SELECT * FROM locations WHERE company_id = ?"),
            (company_id,),
        )
        locations = _rows(cur)
        # 5. departments
        cur = conn.execute(
            q("SELECT * FROM departments WHERE company_id = ?"),
            (company_id,),
        )
        departments = _rows(cur)
        # 6. phone_routing (scope through location)
        cur = conn.execute(
            q("""SELECT phr.* FROM phone_routing phr
                 JOIN locations l ON l.location_id = phr.location_id
                 WHERE l.company_id = ?"""),
            (company_id,),
        )
        phone_routings = _rows(cur)
        # 7. rubric_groups (scope through location)
        cur = conn.execute(
            q("""SELECT rg.* FROM rubric_groups rg
                 JOIN locations l ON l.location_id = rg.location_id
                 WHERE l.company_id = ?"""),
            (company_id,),
        )
        rubric_groups = _rows(cur)
        rubric_group_ids = [r["rubric_group_id"] for r in rubric_groups]
        # 8. rubric_items
        rubric_items = []
        if rubric_group_ids:
            placeholders = ",".join(["?"] * len(rubric_group_ids))
            cur = conn.execute(
                q(f"SELECT * FROM rubric_items WHERE rubric_group_id IN ({placeholders})"),
                rubric_group_ids,
            )
            rubric_items = _rows(cur)
        # 9. projects
        cur = conn.execute(
            q("SELECT * FROM projects WHERE company_id = ?"),
            (company_id,),
        )
        projects = _rows(cur)
        project_ids = [p["project_id"] for p in projects]
        # 10. interactions  (strip binary audio blob)
        interactions = []
        if project_ids:
            placeholders = ",".join(["?"] * len(project_ids))
            cur = conn.execute(
                q(f"""SELECT interaction_id, project_id, caller_user_id,
                             respondent_user_id, interaction_date,
                             interaction_submitted_at, status_id,
                             interaction_transcript, interaction_audio_url,
                             interaction_overall_score,
                             interaction_original_score,
                             interaction_regrade_count,
                             interaction_regraded_with_context,
                             interaction_reviewer_context,
                             interaction_strengths, interaction_weaknesses,
                             interaction_overall_assessment,
                             interaction_flags, interaction_responder_name,
                             interaction_deleted_at, interaction_created_at,
                             interaction_updated_at
                      FROM interactions
                      WHERE project_id IN ({placeholders})"""),
                project_ids,
            )
            interactions = _rows(cur)
        interaction_ids = [i["interaction_id"] for i in interactions]
        # 11. interaction_rubric_scores
        interaction_rubric_scores = []
        clarifying_questions = []
        if interaction_ids:
            placeholders = ",".join(["?"] * len(interaction_ids))
            cur = conn.execute(
                q(f"""SELECT * FROM interaction_rubric_scores
                      WHERE interaction_id IN ({placeholders})"""),
                interaction_ids,
            )
            interaction_rubric_scores = _rows(cur)
            cur = conn.execute(
                q(f"""SELECT * FROM clarifying_questions
                      WHERE interaction_id IN ({placeholders})"""),
                interaction_ids,
            )
            clarifying_questions = _rows(cur)
        # 12. performance_reports (scope through user → department)
        cur = conn.execute(
            q("""SELECT pr.* FROM performance_reports pr
                 JOIN users u       ON u.user_id = pr.subject_user_id
                 JOIN departments d ON d.department_id = u.department_id
                 WHERE d.company_id = ?"""),
            (company_id,),
        )
        performance_reports = _rows(cur)
        # 13. voip_call_queue (strip recording_data)
        cur = conn.execute(
            q("""SELECT voip_queue_id, company_id, voip_queue_provider,
                        voip_queue_call_id, voip_queue_recording_url,
                        voip_queue_caller_number, voip_queue_called_number,
                        voip_queue_call_date, voip_queue_duration_seconds,
                        voip_queue_raw_payload, voip_queue_status,
                        voip_queue_error, voip_queue_interaction_id,
                        voip_queue_created_at, voip_queue_updated_at
                 FROM voip_call_queue WHERE company_id = ?"""),
            (company_id,),
        )
        voip_call_queue = _rows(cur)
        # 14. users (sanitized — no password hashes)
        cur = conn.execute(
            q(f"""SELECT {_SANITIZED_USER_COLUMNS}
                  FROM users u
                  JOIN departments d ON d.department_id = u.department_id
                  WHERE d.company_id = ?"""),
            (company_id,),
        )
        users = _rows(cur)
    finally:
        conn.close()

    payload = {
        "version":    "echo-audit-v2-backup-1",
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "company_id": company_id,
        "tables": {
            "companies":                 companies,
            "company_settings":          settings,
            "company_labels":            labels,
            "locations":                 locations,
            "departments":               departments,
            "phone_routings":            phone_routings,
            "rubric_groups":             rubric_groups,
            "rubric_items":              rubric_items,
            "projects":                  projects,
            "interactions":              interactions,
            "interaction_rubric_scores": interaction_rubric_scores,
            "clarifying_questions":      clarifying_questions,
            "performance_reports":       performance_reports,
            "voip_call_queue":           voip_call_queue,
            "users":                     users,
        },
    }

    # Audit
    write_audit_log(
        current_user.user_id, ACTION_EXPORTED, ENTITY_COMPANY, company_id,
        metadata={"action": "export_backup",
                  "counts": {k: len(v) for k, v in payload["tables"].items()}},
    )

    blob = json.dumps(payload, default=_json_default, indent=2).encode("utf-8")
    filename = f"echoaudit_backup_{date.today().isoformat()}.json"
    return send_file(
        io.BytesIO(blob),
        mimetype="application/json",
        as_attachment=True,
        download_name=filename,
    )


# ═══════════════════════════════════════════════════════════════
# POST /api/export/restore  — JSON restore (all-or-nothing)
# ═══════════════════════════════════════════════════════════════
#
# Restore strategy:
# - All writes happen inside a single DB transaction; on any exception
#   the whole thing rolls back via a single commit/rollback in the route.
# - Original PKs are mapped to newly issued PKs so relationships are
#   preserved without requiring identity-insert. Every FK is remapped
#   through `id_maps`.
# - Restore is scoped to the CURRENT company context — the backup's own
#   `company_id` is overridden to whatever the caller is operating as.
# - Users are created with user_must_change_password = TRUE and a random
#   disabled-looking password hash so restored accounts can never log in
#   on the restored side until the admin resets passwords via the platform
#   reset-password surface.
# - interaction_audio_data, voip_queue_recording_data, voip_configs, and
#   audit_log are NOT restored (they're not in the backup either).


_TABLE_ORDER = (
    "locations",
    "departments",
    "users",
    "phone_routings",
    "rubric_groups",
    "rubric_items",
    "projects",
    "interactions",
    "interaction_rubric_scores",
    "clarifying_questions",
    "performance_reports",
    "voip_call_queue",
    "company_settings",
    "company_labels",
)


def _random_placeholder_hash():
    """A deliberately non-verifying password hash for restored users.
    Nothing checks out against this; any subsequent login will fail until
    an admin runs /api/platform/users/<id>/reset-password.
    """
    import secrets as _sec
    return "pbkdf2:sha256:260000$restored$" + _sec.token_hex(32)


def _validate_backup_structure(payload):
    if not isinstance(payload, dict):
        return "Backup must be a JSON object"
    if payload.get("version") != "echo-audit-v2-backup-1":
        return "Unsupported backup version"
    tables = payload.get("tables")
    if not isinstance(tables, dict):
        return "'tables' must be an object"
    for t in _TABLE_ORDER:
        if t not in tables:
            return f"Backup is missing table: {t}"
        if not isinstance(tables[t], list):
            return f"Backup table '{t}' must be an array"
    return None


def _insert_row_returning(conn, table, row, pk_col):
    """Insert a row dict and return the new PK. Columns in `row` drive the
    INSERT shape so callers can omit server-managed fields like auto PKs and
    timestamps by removing them from the dict before calling.
    """
    cols = list(row.keys())
    if IS_POSTGRES:
        placeholders = ", ".join(["%s"] * len(cols))
        sql = (f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({placeholders}) "
               f"RETURNING {pk_col}")
        cur = conn.execute(sql, [row[c] for c in cols])
        return cur.fetchone()[pk_col]
    placeholders = ", ".join(["?"] * len(cols))
    conn.execute(
        f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({placeholders})",
        [row[c] for c in cols],
    )
    return conn.execute("SELECT last_insert_rowid()").fetchone()[0]


def _pop_managed(row, *cols):
    """Remove server-managed columns from a restore row before INSERT."""
    for c in cols:
        row.pop(c, None)
    return row


def _remap_fk(row, field, mapping):
    """Translate an old-ID FK to a new one. Sets to None if no mapping exists."""
    val = row.get(field)
    if val is None:
        return
    row[field] = mapping.get(val)


def _dumps_jsonb(value):
    """Normalize JSONB values for insertion across backends."""
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    return value


@export_bp.route("/restore", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def restore_backup():
    company_id, err = _require_company()
    if err: return err

    # Accept either multipart upload or raw JSON body
    payload = None
    if request.files and "file" in request.files:
        try:
            payload = json.loads(request.files["file"].read().decode("utf-8"))
        except Exception:
            return _err("Uploaded file is not valid JSON", 400)
    else:
        payload = request.get_json(silent=True)
        if payload is None:
            return _err("Request body must be JSON or multipart with 'file'", 400)

    structure_err = _validate_backup_structure(payload)
    if structure_err:
        return _err(structure_err, 400)

    tables = payload["tables"]

    # id_maps[table][old_id] = new_id
    id_maps: dict[str, dict] = {t: {} for t in _TABLE_ORDER}

    counts = {}

    conn = get_conn()
    try:
        # Stay in a single txn: ensure autocommit is off on Postgres.
        if IS_POSTGRES:
            try:
                conn.autocommit = False
            except Exception:
                pass

        # ── 1. locations ──
        for row in tables["locations"]:
            row = dict(row)
            old_id = row.pop("location_id", None)
            row["company_id"] = company_id
            _pop_managed(row, "location_created_at", "location_updated_at")
            new_id = _insert_row_returning(conn, "locations", row, "location_id")
            if old_id is not None:
                id_maps["locations"][old_id] = new_id
        counts["locations"] = len(id_maps["locations"])

        # ── 2. departments ──
        for row in tables["departments"]:
            row = dict(row)
            old_id = row.pop("department_id", None)
            row["company_id"] = company_id
            _pop_managed(row, "department_created_at", "department_updated_at")
            new_id = _insert_row_returning(conn, "departments", row, "department_id")
            if old_id is not None:
                id_maps["departments"][old_id] = new_id
        counts["departments"] = len(id_maps["departments"])

        # ── 3. users (sanitized) ──
        for row in tables["users"]:
            row = dict(row)
            old_id = row.pop("user_id", None)
            _remap_fk(row, "department_id", id_maps["departments"])
            _pop_managed(row, "user_created_at", "user_updated_at")
            # Users without a remapped department can't be restored; skip silently.
            if row.get("department_id") is None and "department_id" in row:
                continue
            # Force password reset on restored users and install a non-verifying
            # placeholder hash — never carries a real password across.
            row["user_password_hash"] = _random_placeholder_hash()
            row["user_must_change_password"] = True
            new_id = _insert_row_returning(conn, "users", row, "user_id")
            if old_id is not None:
                id_maps["users"][old_id] = new_id
        counts["users"] = len(id_maps["users"])

        # ── 4. phone_routings ──
        for row in tables["phone_routings"]:
            row = dict(row)
            old_id = row.pop("phone_routing_id", None)
            _remap_fk(row, "location_id", id_maps["locations"])
            if row.get("location_id") is None:
                continue
            _pop_managed(row, "phone_routing_created_at", "phone_routing_updated_at")
            new_id = _insert_row_returning(conn, "phone_routing", row, "phone_routing_id")
            if old_id is not None:
                id_maps["phone_routings"][old_id] = new_id
        counts["phone_routings"] = len(id_maps["phone_routings"])

        # ── 5. rubric_groups ──
        for row in tables["rubric_groups"]:
            row = dict(row)
            old_id = row.pop("rubric_group_id", None)
            _remap_fk(row, "location_id", id_maps["locations"])
            if row.get("location_id") is None:
                continue
            _pop_managed(row, "rg_created_at", "rg_updated_at")
            new_id = _insert_row_returning(conn, "rubric_groups", row, "rubric_group_id")
            if old_id is not None:
                id_maps["rubric_groups"][old_id] = new_id
        counts["rubric_groups"] = len(id_maps["rubric_groups"])

        # ── 6. rubric_items ──
        for row in tables["rubric_items"]:
            row = dict(row)
            old_id = row.pop("rubric_item_id", None)
            _remap_fk(row, "rubric_group_id", id_maps["rubric_groups"])
            if row.get("rubric_group_id") is None:
                continue
            _pop_managed(row, "ri_created_at", "ri_updated_at")
            new_id = _insert_row_returning(conn, "rubric_items", row, "rubric_item_id")
            if old_id is not None:
                id_maps["rubric_items"][old_id] = new_id
        counts["rubric_items"] = len(id_maps["rubric_items"])

        # ── 7. projects ──
        for row in tables["projects"]:
            row = dict(row)
            old_id = row.pop("project_id", None)
            row["company_id"] = company_id
            _remap_fk(row, "phone_routing_id", id_maps["phone_routings"])
            _remap_fk(row, "rubric_group_id", id_maps["rubric_groups"])
            if row.get("rubric_group_id") is None:
                continue
            _pop_managed(row, "project_created_at", "project_updated_at")
            new_id = _insert_row_returning(conn, "projects", row, "project_id")
            if old_id is not None:
                id_maps["projects"][old_id] = new_id
        counts["projects"] = len(id_maps["projects"])

        # ── 8. interactions ──
        for row in tables["interactions"]:
            row = dict(row)
            old_id = row.pop("interaction_id", None)
            _remap_fk(row, "project_id", id_maps["projects"])
            _remap_fk(row, "caller_user_id", id_maps["users"])
            _remap_fk(row, "respondent_user_id", id_maps["users"])
            _pop_managed(row, "interaction_created_at", "interaction_updated_at")
            # Audio blob is not in the backup; leave NULL.
            row.setdefault("interaction_audio_data", None)
            new_id = _insert_row_returning(conn, "interactions", row, "interaction_id")
            if old_id is not None:
                id_maps["interactions"][old_id] = new_id
        counts["interactions"] = len(id_maps["interactions"])

        # ── 9. interaction_rubric_scores ──
        restored_irs = 0
        for row in tables["interaction_rubric_scores"]:
            row = dict(row)
            row.pop("interaction_rubric_score_id", None)
            _remap_fk(row, "interaction_id", id_maps["interactions"])
            _remap_fk(row, "rubric_item_id", id_maps["rubric_items"])
            if row.get("interaction_id") is None:
                continue
            _pop_managed(row, "irs_created_at", "irs_updated_at")
            _insert_row_returning(conn, "interaction_rubric_scores", row,
                                  "interaction_rubric_score_id")
            restored_irs += 1
        counts["interaction_rubric_scores"] = restored_irs

        # ── 10. clarifying_questions ──
        restored_cqs = 0
        for row in tables["clarifying_questions"]:
            row = dict(row)
            row.pop("clarifying_question_id", None)
            _remap_fk(row, "interaction_id", id_maps["interactions"])
            if row.get("interaction_id") is None:
                continue
            _pop_managed(row, "cq_created_at", "cq_updated_at")
            _insert_row_returning(conn, "clarifying_questions", row,
                                  "clarifying_question_id")
            restored_cqs += 1
        counts["clarifying_questions"] = restored_cqs

        # ── 11. performance_reports ──
        restored_pr = 0
        for row in tables["performance_reports"]:
            row = dict(row)
            row.pop("performance_report_id", None)
            _remap_fk(row, "subject_user_id", id_maps["users"])
            if row.get("subject_user_id") is None:
                continue
            _pop_managed(row, "pr_created_at", "pr_updated_at")
            # JSONB fields need string form for our helpers
            row["pr_data"] = _dumps_jsonb(row.get("pr_data"))
            row["pr_processed_interaction_ids"] = _dumps_jsonb(
                row.get("pr_processed_interaction_ids")
            )
            _insert_row_returning(conn, "performance_reports", row,
                                  "performance_report_id")
            restored_pr += 1
        counts["performance_reports"] = restored_pr

        # ── 12. voip_call_queue ──
        restored_vq = 0
        for row in tables["voip_call_queue"]:
            row = dict(row)
            row.pop("voip_queue_id", None)
            row["company_id"] = company_id
            _remap_fk(row, "voip_queue_interaction_id", id_maps["interactions"])
            _pop_managed(row, "voip_queue_created_at", "voip_queue_updated_at")
            row["voip_queue_raw_payload"] = _dumps_jsonb(row.get("voip_queue_raw_payload"))
            # Recording blob isn't in the backup — leave NULL.
            row.setdefault("voip_queue_recording_data", None)
            _insert_row_returning(conn, "voip_call_queue", row, "voip_queue_id")
            restored_vq += 1
        counts["voip_call_queue"] = restored_vq

        # ── 13. company_settings ── upsert rather than blind insert so the
        # per-company defaults seeded at company creation don't collide.
        restored_settings = 0
        for row in tables["company_settings"]:
            row = dict(row)
            row.pop("company_setting_id", None)
            row["company_id"] = company_id
            _pop_managed(row, "company_setting_updated_at")
            key = row.get("company_setting_key")
            value = row.get("company_setting_value")
            if not key or value is None:
                continue
            if IS_POSTGRES:
                conn.execute(
                    """INSERT INTO company_settings
                           (company_id, company_setting_key, company_setting_value)
                       VALUES (%s, %s, %s)
                       ON CONFLICT (company_id, company_setting_key) DO UPDATE
                       SET company_setting_value      = EXCLUDED.company_setting_value,
                           company_setting_updated_at = NOW()""",
                    (company_id, key, value),
                )
            else:
                conn.execute(
                    """INSERT INTO company_settings
                           (company_id, company_setting_key, company_setting_value)
                       VALUES (?, ?, ?)
                       ON CONFLICT (company_id, company_setting_key) DO UPDATE
                       SET company_setting_value = excluded.company_setting_value,
                           company_setting_updated_at = CURRENT_TIMESTAMP""",
                    (company_id, key, value),
                )
            restored_settings += 1
        counts["company_settings"] = restored_settings

        # ── 14. company_labels ── upsert
        restored_labels = 0
        for row in tables["company_labels"]:
            row = dict(row)
            row.pop("company_label_id", None)
            row["company_id"] = company_id
            _pop_managed(row, "cl_created_at", "cl_updated_at")
            key = row.get("cl_key")
            value = row.get("cl_value")
            if not key or value is None:
                continue
            if IS_POSTGRES:
                conn.execute(
                    """INSERT INTO company_labels (company_id, cl_key, cl_value)
                       VALUES (%s, %s, %s)
                       ON CONFLICT (company_id, cl_key) DO UPDATE
                       SET cl_value = EXCLUDED.cl_value, cl_updated_at = NOW()""",
                    (company_id, key, value),
                )
            else:
                conn.execute(
                    """INSERT INTO company_labels (company_id, cl_key, cl_value)
                       VALUES (?, ?, ?)
                       ON CONFLICT (company_id, cl_key) DO UPDATE
                       SET cl_value = excluded.cl_value,
                           cl_updated_at = CURRENT_TIMESTAMP""",
                    (company_id, key, value),
                )
            restored_labels += 1
        counts["company_labels"] = restored_labels

        # Audit inside the same txn so it rolls back too if something explodes.
        write_audit_log(
            current_user.user_id, ACTION_UPDATED, ENTITY_COMPANY, company_id,
            metadata={"action": "backup_restored", "counts": counts},
            conn=conn,
        )

        conn.commit()
    except Exception as exc:
        conn.rollback()
        logger.exception("Restore failed; rolled back")
        return _err(f"Restore failed: {exc}", 400)
    finally:
        conn.close()

    return jsonify({"ok": True, "counts": counts})
