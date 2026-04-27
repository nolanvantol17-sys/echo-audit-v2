"""
api_routes.py — Echo Audit V2 Phase 2 API routes.

Scope: companies, locations, departments, phone_routing, projects,
       team (users), and industries reference data.

All routes emit JSON. Company scoping is enforced via
get_effective_company_id() on every route except /api/companies
(super-admin cross-tenant) and /api/industries (global reference).
"""

import logging
from datetime import date, timedelta

from flask import Blueprint, jsonify, request
from flask_login import current_user, login_required

import auth
from audit_log import (
    ACTION_CREATED, ACTION_DELETED, ACTION_UPDATED,
    ENTITY_CAMPAIGN, ENTITY_COMPANY, ENTITY_DEPARTMENT, ENTITY_LOCATION,
    ENTITY_PHONE_ROUTING, ENTITY_PROJECT, ENTITY_USER,
    write_audit_log,
)
from auth import role_required
from dashboard_helpers import (
    _month_bounds, _report_url_for, _roll_up_locations, _trend_for_calls,
)
from db import get_conn, q, IS_POSTGRES
from helpers import generate_temp_password, get_effective_company_id

logger = logging.getLogger(__name__)

api_bp = Blueprint("api", __name__, url_prefix="/api")


# ── Response helpers ────────────────────────────────────────────


def _err(msg, code):
    return jsonify({"error": msg}), code


def _ok():
    return jsonify({"ok": True})


def _body():
    """Return JSON body as a dict, or {} if none / malformed."""
    return request.get_json(silent=True) or {}


def _require(body, *fields):
    """Return None if all fields present and non-blank, else an error string."""
    missing = [f for f in fields if body.get(f) in (None, "")]
    if missing:
        return f"Missing required fields: {', '.join(missing)}"
    return None


def _require_company():
    """Resolve current company scope or return (None, error_response)."""
    cid = get_effective_company_id()
    if cid is None:
        return None, _err(
            "No company context. Super admins must select an organization first.",
            400,
        )
    return cid, None


def _row_to_dict(row):
    """Normalize psycopg2 RealDictRow / sqlite3.Row to a plain dict."""
    if row is None:
        return None
    try:
        return dict(row)
    except Exception:
        return {k: row[k] for k in row.keys()}  # sqlite3.Row fallback


def _rows(cur):
    return [_row_to_dict(r) for r in cur.fetchall()]


# ── Ownership / fetch helpers ───────────────────────────────────


def _get_company(conn, company_id):
    cur = conn.execute(
        q("SELECT * FROM companies WHERE company_id = ? AND company_deleted_at IS NULL"),
        (company_id,),
    )
    return cur.fetchone()


def _get_location(conn, location_id, company_id):
    cur = conn.execute(
        q("""SELECT * FROM locations
             WHERE location_id = ? AND company_id = ? AND location_deleted_at IS NULL"""),
        (location_id, company_id),
    )
    return cur.fetchone()


def _get_department(conn, department_id, company_id):
    cur = conn.execute(
        q("""SELECT * FROM departments
             WHERE department_id = ? AND company_id = ? AND department_deleted_at IS NULL"""),
        (department_id, company_id),
    )
    return cur.fetchone()


def _get_phone_routing(conn, phone_routing_id, company_id):
    cur = conn.execute(
        q("""SELECT phr.* FROM phone_routing phr
             JOIN locations l ON l.location_id = phr.location_id
             WHERE phr.phone_routing_id = ? AND l.company_id = ?
               AND l.location_deleted_at IS NULL"""),
        (phone_routing_id, company_id),
    )
    return cur.fetchone()


def _get_project(conn, project_id, company_id):
    cur = conn.execute(
        q("""SELECT * FROM projects
             WHERE project_id = ? AND company_id = ? AND project_deleted_at IS NULL"""),
        (project_id, company_id),
    )
    return cur.fetchone()


def _get_campaign_in_company(conn, campaign_id, company_id):
    """Return campaign row if its project belongs to this company (soft-delete aware)."""
    cur = conn.execute(
        q("""SELECT c.* FROM campaigns c
             JOIN projects p ON p.project_id = c.project_id
             WHERE c.campaign_id = ? AND p.company_id = ?
               AND c.campaign_deleted_at IS NULL
               AND p.project_deleted_at IS NULL"""),
        (campaign_id, company_id),
    )
    return cur.fetchone()


def _get_rubric_group_in_company(conn, rubric_group_id, company_id):
    """Rubric group is 'in' a company if its location_id → locations.company_id matches.
    Industry templates (location_id IS NULL) are NOT usable directly."""
    cur = conn.execute(
        q("""SELECT rg.rubric_group_id FROM rubric_groups rg
             JOIN locations l ON l.location_id = rg.location_id
             WHERE rg.rubric_group_id = ? AND l.company_id = ?
               AND rg.rg_deleted_at IS NULL
               AND l.location_deleted_at IS NULL"""),
        (rubric_group_id, company_id),
    )
    return cur.fetchone() is not None


def _get_user_in_company(conn, user_id, company_id):
    cur = conn.execute(
        q("""SELECT u.* FROM users u
             JOIN departments d ON d.department_id = u.department_id
             WHERE u.user_id = ? AND d.company_id = ? AND u.user_deleted_at IS NULL"""),
        (user_id, company_id),
    )
    return cur.fetchone()


def _lastrowid(conn, returning_col=None, cur=None):
    """Return the last inserted ID. Handles psycopg2 RETURNING and sqlite3."""
    if IS_POSTGRES:
        row = cur.fetchone()
        return row[returning_col]
    return conn.execute("SELECT last_insert_rowid()").fetchone()[0]


def _insert_returning(conn, sql_pg, sql_lite, params, pk_col):
    """Helper for INSERTs that must return a generated PK across both backends."""
    if IS_POSTGRES:
        cur = conn.execute(sql_pg, params)
        return cur.fetchone()[pk_col]
    else:
        conn.execute(sql_lite, params)
        return conn.execute("SELECT last_insert_rowid()").fetchone()[0]


# ═══════════════════════════════════════════════════════════════
# COMPANIES  (super_admin only, cross-tenant)
# ═══════════════════════════════════════════════════════════════


@api_bp.route("/companies", methods=["GET"])
@login_required
@role_required("super_admin")
def list_companies():
    conn = get_conn()
    try:
        cur = conn.execute(q("""
            SELECT c.company_id, c.company_name, c.status_id, c.company_created_at,
                   s.status_name
            FROM companies c
            LEFT JOIN statuses s ON s.status_id = c.status_id
            WHERE c.company_deleted_at IS NULL
            ORDER BY c.company_name
        """))
        return jsonify(_rows(cur))
    finally:
        conn.close()


@api_bp.route("/companies", methods=["POST"])
@login_required
@role_required("super_admin")
def create_company_route():
    body = _body()
    err = _require(body, "company_name", "industry_id")
    if err:
        return _err(err, 400)

    conn = get_conn()
    try:
        try:
            company_id = _insert_returning(
                conn,
                sql_pg="""INSERT INTO companies (company_name, industry_id, status_id)
                          VALUES (%s, %s, 1) RETURNING company_id""",
                sql_lite="INSERT INTO companies (company_name, industry_id, status_id) VALUES (?, ?, 1)",
                params=(body["company_name"], body["industry_id"]),
                pk_col="company_id",
            )
            write_audit_log(
                current_user.user_id, ACTION_CREATED, ENTITY_COMPANY, company_id,
                metadata={"company_name": body["company_name"],
                          "industry_id": body["industry_id"]},
                conn=conn,
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            msg = str(e).lower()
            if "duplicate" in msg or "unique" in msg:
                return _err("Company already exists", 409)
            if "foreign key" in msg or "violates foreign key" in msg:
                return _err("Invalid industry_id", 400)
            raise
        row = _get_company(conn, company_id)
        return jsonify(_row_to_dict(row)), 201
    finally:
        conn.close()


@api_bp.route("/companies/<int:company_id>", methods=["PUT"])
@login_required
@role_required("super_admin")
def update_company(company_id):
    body = _body()
    allowed = {"company_name", "industry_id", "status_id"}
    fields = {k: body[k] for k in allowed if k in body}
    if not fields:
        return _err("No fields to update", 400)

    conn = get_conn()
    try:
        if not _get_company(conn, company_id):
            return _err("Company not found", 404)

        sets = ", ".join(f"{k} = ?" for k in fields)
        params = list(fields.values()) + [company_id]
        try:
            conn.execute(q(f"UPDATE companies SET {sets} WHERE company_id = ?"), params)
            write_audit_log(
                current_user.user_id, ACTION_UPDATED, ENTITY_COMPANY, company_id,
                metadata={"changes": fields}, conn=conn,
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            if "foreign key" in str(e).lower():
                return _err("Invalid industry_id or status_id", 400)
            raise

        return jsonify(_row_to_dict(_get_company(conn, company_id)))
    finally:
        conn.close()


@api_bp.route("/companies/<int:company_id>/deactivate", methods=["POST"])
@login_required
@role_required("super_admin")
def deactivate_company(company_id):
    conn = get_conn()
    try:
        if not _get_company(conn, company_id):
            return _err("Company not found", 404)
        conn.execute(q("UPDATE companies SET status_id = 10 WHERE company_id = ?"),
                     (company_id,))
        write_audit_log(
            current_user.user_id, ACTION_DELETED, ENTITY_COMPANY, company_id,
            metadata={"action": "deactivate", "new_status_id": 10}, conn=conn,
        )
        conn.commit()
        return _ok()
    finally:
        conn.close()


@api_bp.route("/companies/<int:company_id>/reactivate", methods=["POST"])
@login_required
@role_required("super_admin")
def reactivate_company(company_id):
    conn = get_conn()
    try:
        if not _get_company(conn, company_id):
            return _err("Company not found", 404)
        conn.execute(q("UPDATE companies SET status_id = 1 WHERE company_id = ?"),
                     (company_id,))
        write_audit_log(
            current_user.user_id, ACTION_UPDATED, ENTITY_COMPANY, company_id,
            metadata={"action": "reactivate", "new_status_id": 1}, conn=conn,
        )
        conn.commit()
        return _ok()
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
# LOCATIONS
# ═══════════════════════════════════════════════════════════════


@api_bp.route("/locations", methods=["GET"])
@login_required
@role_required("admin", "super_admin")
def list_locations():
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        cur = conn.execute(q("""
            SELECT l.location_id, l.location_name, l.location_phone,
                   l.location_engagement_date, l.status_id, s.status_name,
                   COALESCE(m.total_calls, 0)     AS total_calls,
                   COALESCE(m.graded_count, 0)    AS graded_count,
                   COALESCE(m.no_answer_count, 0) AS no_answer_count,
                   m.avg_score                    AS avg_score,
                   m.last_call_date               AS last_call_date
            FROM locations l
            LEFT JOIN statuses s ON s.status_id = l.status_id
            LEFT JOIN (
                SELECT i.interaction_location_id AS location_id,
                       COUNT(*)                                                  AS total_calls,
                       SUM(CASE WHEN i.status_id = 43 THEN 1 ELSE 0 END)         AS graded_count,
                       SUM(CASE WHEN i.status_id = 44 THEN 1 ELSE 0 END)         AS no_answer_count,
                       AVG(CASE WHEN i.status_id = 43
                                THEN i.interaction_overall_score END)            AS avg_score,
                       MAX(CASE WHEN i.status_id = 43
                                THEN i.interaction_date END)                     AS last_call_date
                  FROM interactions i
                  JOIN projects p ON p.project_id = i.project_id
                 WHERE p.company_id = ?
                   AND i.interaction_deleted_at IS NULL
                 GROUP BY i.interaction_location_id
            ) m ON m.location_id = l.location_id
            WHERE l.company_id = ? AND l.location_deleted_at IS NULL
            ORDER BY l.location_name
        """), (company_id, company_id))
        rows = _rows(cur)
        # AVG returns Decimal in PG; coerce to JSON-safe float (matches the
        # round(float(...), 1) pattern used elsewhere in this file).
        # SUM/COUNT can also return Decimal in PG — coerce to int.
        for r in rows:
            avg = r.get("avg_score")
            r["avg_score"] = round(float(avg), 1) if avg is not None else None
            for k in ("total_calls", "graded_count", "no_answer_count"):
                v = r.get(k)
                if v is not None:
                    r[k] = int(v)
            # Derived: no_answer_rate = no_ans / (graded + no_ans). Null when
            # the denominator is zero (no terminal calls yet for this loc).
            g = r.get("graded_count") or 0
            n = r.get("no_answer_count") or 0
            r["no_answer_rate"] = (n / (g + n)) if (g + n) else None
        return jsonify(rows)
    finally:
        conn.close()


@api_bp.route("/locations", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def create_location():
    company_id, err = _require_company()
    if err: return err

    body = _body()
    err = _require(body, "location_name")
    if err:
        return _err(err, 400)

    conn = get_conn()
    try:
        try:
            location_id = _insert_returning(
                conn,
                sql_pg="""INSERT INTO locations
                              (company_id, location_name, location_phone,
                               location_engagement_date, status_id)
                          VALUES (%s, %s, %s, %s, 1) RETURNING location_id""",
                sql_lite="""INSERT INTO locations
                                (company_id, location_name, location_phone,
                                 location_engagement_date, status_id)
                            VALUES (?, ?, ?, ?, 1)""",
                params=(
                    company_id,
                    body["location_name"],
                    body.get("location_phone"),
                    body.get("location_engagement_date"),
                ),
                pk_col="location_id",
            )
            write_audit_log(
                current_user.user_id, ACTION_CREATED, ENTITY_LOCATION, location_id,
                metadata={"company_id": company_id,
                          "location_name": body["location_name"]},
                conn=conn,
            )
            conn.commit()
        except Exception:
            conn.rollback()
            raise

        return jsonify(_row_to_dict(_get_location(conn, location_id, company_id))), 201
    finally:
        conn.close()


@api_bp.route("/locations/<int:location_id>", methods=["PUT"])
@login_required
@role_required("admin", "super_admin")
def update_location(location_id):
    company_id, err = _require_company()
    if err: return err

    body = _body()
    allowed = {"location_name", "location_phone", "location_engagement_date", "status_id"}
    fields = {k: body[k] for k in allowed if k in body}
    if not fields:
        return _err("No fields to update", 400)

    conn = get_conn()
    try:
        if not _get_location(conn, location_id, company_id):
            return _err("Location not found", 404)

        sets = ", ".join(f"{k} = ?" for k in fields)
        params = list(fields.values()) + [location_id]
        conn.execute(q(f"UPDATE locations SET {sets} WHERE location_id = ?"), params)
        write_audit_log(
            current_user.user_id, ACTION_UPDATED, ENTITY_LOCATION, location_id,
            metadata={"changes": fields}, conn=conn,
        )
        conn.commit()
        return jsonify(_row_to_dict(_get_location(conn, location_id, company_id)))
    finally:
        conn.close()


@api_bp.route("/locations/<int:location_id>", methods=["DELETE"])
@login_required
@role_required("admin", "super_admin")
def delete_location(location_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        if not _get_location(conn, location_id, company_id):
            return _err("Location not found", 404)
        conn.execute(q("""UPDATE locations SET location_deleted_at = CURRENT_TIMESTAMP
                          WHERE location_id = ?"""), (location_id,))
        write_audit_log(
            current_user.user_id, ACTION_DELETED, ENTITY_LOCATION, location_id,
            conn=conn,
        )
        conn.commit()
        return _ok()
    finally:
        conn.close()


@api_bp.route("/locations/<int:location_id>/deletion-impact", methods=["GET"])
@login_required
@role_required("admin", "super_admin")
def location_deletion_impact(location_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        loc = _get_location(conn, location_id, company_id)
        if not loc:
            return _err("Location not found", 404)

        cur = conn.execute(q("""
            SELECT COUNT(*) AS cnt FROM rubric_groups
            WHERE location_id = ? AND rg_deleted_at IS NULL
        """), (location_id,))
        row = cur.fetchone()
        rubrics_count = row["cnt"] if IS_POSTGRES else row[0]

        cur = conn.execute(q("""
            SELECT COUNT(DISTINCT p.project_id) AS cnt FROM projects p
            LEFT JOIN phone_routing phr ON phr.phone_routing_id = p.phone_routing_id
            LEFT JOIN rubric_groups rg ON rg.rubric_group_id = p.rubric_group_id
            WHERE p.company_id = ? AND p.project_deleted_at IS NULL
              AND (phr.location_id = ? OR rg.location_id = ?)
        """), (company_id, location_id, location_id))
        row = cur.fetchone()
        projects_count = row["cnt"] if IS_POSTGRES else row[0]

        return jsonify({
            "deletable": True,
            "name": dict(loc).get("location_name"),
            "counts": {"rubrics": rubrics_count, "projects": projects_count},
        })
    finally:
        conn.close()


# ── Bulk upload ────────────────────────────────────────────────
# Accepts CSV / Excel. Flexible column matching, tolerant of extra
# columns and blank rows. Duplicates (by case-insensitive name within
# the current company) are flagged in `warnings` but still have their
# phone / engagement_date updated from the file. Per-row errors never
# abort the batch.

_NAME_ALIASES  = {"name", "locationname", "propertyname"}
_PHONE_ALIASES = {"phone", "phonenumber", "telephone", "locationphone"}
_DATE_ALIASES  = {"engagementdate", "date", "locationengagementdate"}


def _norm_header(h):
    return "".join(ch for ch in str(h or "").lower() if ch.isalnum())


def _map_headers(headers):
    """Return {canonical_key: source_index} for recognized columns."""
    mapping = {}
    for idx, h in enumerate(headers):
        key = _norm_header(h)
        if key in _NAME_ALIASES and "location_name" not in mapping:
            mapping["location_name"] = idx
        elif key in _PHONE_ALIASES and "location_phone" not in mapping:
            mapping["location_phone"] = idx
        elif key in _DATE_ALIASES and "location_engagement_date" not in mapping:
            mapping["location_engagement_date"] = idx
    return mapping


def _parse_engagement_date(value):
    """Best-effort date parse. Returns 'YYYY-MM-DD' or None."""
    import datetime as _dt
    if value in (None, ""):
        return None
    if isinstance(value, _dt.datetime):
        return value.date().isoformat()
    if isinstance(value, _dt.date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y",
                "%Y/%m/%d", "%b %d, %Y", "%B %d, %Y"):
        try:
            return _dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s  # let DB reject if truly unparseable


def _read_csv_rows(stream):
    import csv, io
    raw = stream.read()
    if isinstance(raw, bytes):
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw.decode("utf-8", errors="replace")
    else:
        text = raw
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _read_xlsx_rows(stream):
    from openpyxl import load_workbook
    wb = load_workbook(stream, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    data_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
        else:
            data_rows.append(list(row))
    return headers, data_rows


@api_bp.route("/locations/bulk-upload", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def bulk_upload_locations():
    company_id, err = _require_company()
    if err: return err

    f = request.files.get("file")
    if not f or not f.filename:
        return _err("No file uploaded.", 400)

    name = f.filename.lower()
    if name.endswith(".csv"):
        headers, data_rows = _read_csv_rows(f.stream)
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        try:
            headers, data_rows = _read_xlsx_rows(f.stream)
        except Exception as e:
            logger.exception("bulk_upload_locations: xlsx parse failed")
            return _err("Could not read Excel file: " + str(e), 400)
    else:
        return _err("Unsupported file type. Please upload a CSV or Excel file.", 400)

    col_map = _map_headers(headers or [])
    if "location_name" not in col_map:
        return _err("File must include a 'Name' column.", 400)

    def cell(row, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return "" if v is None else str(v).strip()

    conn = get_conn()
    created = updated = skipped = 0
    warnings_out = []
    errors_out = []

    try:
        # Preload existing names for case-insensitive dup check.
        cur = conn.execute(q("""
            SELECT location_id, location_name FROM locations
            WHERE company_id = ? AND location_deleted_at IS NULL
        """), (company_id,))
        existing = {}
        for r in cur.fetchall():
            d = _row_to_dict(r)
            existing[d["location_name"].strip().lower()] = d["location_id"]

        try:
            for i, raw in enumerate(data_rows):
                row_num = i + 2  # 1-based + header

                if not raw or all((v is None or str(v).strip() == "") for v in raw):
                    continue  # silent skip

                loc_name  = cell(raw, "location_name")
                loc_phone = cell(raw, "location_phone") or None
                loc_date  = _parse_engagement_date(cell(raw, "location_engagement_date")) or None

                if not loc_name:
                    errors_out.append({"row": row_num, "message": "Location name is required"})
                    skipped += 1
                    continue

                dup_id = existing.get(loc_name.lower())
                if dup_id:
                    fields = {}
                    if loc_phone is not None:
                        fields["location_phone"] = loc_phone
                    if loc_date is not None:
                        fields["location_engagement_date"] = loc_date
                    if fields:
                        sets = ", ".join(f"{k} = ?" for k in fields)
                        params = list(fields.values()) + [dup_id]
                        conn.execute(q(f"UPDATE locations SET {sets} WHERE location_id = ?"), params)
                        write_audit_log(
                            current_user.user_id, ACTION_UPDATED, ENTITY_LOCATION, dup_id,
                            metadata={"source": "bulk_upload", "changes": fields}, conn=conn,
                        )
                        updated += 1
                        warnings_out.append({
                            "row": row_num, "location_name": loc_name,
                            "message": "Location already exists — phone number updated",
                        })
                    else:
                        warnings_out.append({
                            "row": row_num, "location_name": loc_name,
                            "message": "Location already exists — no changes",
                        })
                    continue

                try:
                    new_id = _insert_returning(
                        conn,
                        sql_pg="""INSERT INTO locations
                                      (company_id, location_name, location_phone,
                                       location_engagement_date, status_id)
                                  VALUES (%s, %s, %s, %s, 1) RETURNING location_id""",
                        sql_lite="""INSERT INTO locations
                                        (company_id, location_name, location_phone,
                                         location_engagement_date, status_id)
                                    VALUES (?, ?, ?, ?, 1)""",
                        params=(company_id, loc_name, loc_phone, loc_date),
                        pk_col="location_id",
                    )
                    write_audit_log(
                        current_user.user_id, ACTION_CREATED, ENTITY_LOCATION, new_id,
                        metadata={"source": "bulk_upload", "location_name": loc_name},
                        conn=conn,
                    )
                    existing[loc_name.lower()] = new_id
                    created += 1
                except Exception as e:
                    logger.exception("bulk_upload_locations: row %d insert failed", row_num)
                    errors_out.append({"row": row_num, "message": "Insert failed: " + str(e)})
                    skipped += 1

            conn.commit()
        except Exception:
            conn.rollback()
            raise
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "created":  created,
        "updated":  updated,
        "skipped":  skipped,
        "warnings": warnings_out,
        "errors":   errors_out,
    })


# ── Per-location pre-call intel ────────────────────────────────
# Any authenticated user can read the intel for a location in their company.
# Returned by the grade form to populate the property intel card.

# Mirrors intel.py — kept local to avoid a cross-module import for one int.
_STATUS_NO_ANSWER = 44


@api_bp.route("/locations/<int:location_id>/intel", methods=["GET"])
@login_required
def get_location_intel(location_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        # Tenant scope: location must belong to the current company.
        cur = conn.execute(
            q("""SELECT location_id, location_name FROM locations
                  WHERE location_id = ? AND company_id = ?
                    AND location_deleted_at IS NULL"""),
            (location_id, company_id),
        )
        if cur.fetchone() is None:
            return _err("Location not found", 404)

        cur = conn.execute(
            q("""SELECT li_last_call_date, li_last_call_time, li_last_call_score,
                        li_last_call_outcome, li_total_calls, li_avg_score,
                        li_no_answer_count, li_summary, li_strengths, li_weaknesses,
                        li_last_computed_at
                   FROM location_intel
                  WHERE location_id = ? AND company_id = ?"""),
            (location_id, company_id),
        )
        row = _row_to_dict(cur.fetchone())

        # Recent no-answers — pure SQL list, surfaced alongside the AI brief.
        # Helps callers spot bad-time-of-day patterns ("3 dead calls at 2pm").
        # NULLS LAST is PG-only; SQLite dev env will sort NULLs first which is
        # acceptable for local testing only.
        cur = conn.execute(
            q("""SELECT i.interaction_id,
                        i.interaction_date,
                        i.interaction_call_start_time,
                        i.interaction_uploaded_at,
                        u.user_first_name,
                        u.user_last_name
                   FROM interactions i
                   LEFT JOIN users u ON u.user_id = i.caller_user_id
                  WHERE i.interaction_location_id = ?
                    AND i.status_id = ?
                    AND i.interaction_deleted_at IS NULL
                  ORDER BY COALESCE(i.interaction_call_start_time,
                                    i.interaction_uploaded_at) DESC,
                           i.interaction_id DESC
                  LIMIT 5"""),
            (location_id, _STATUS_NO_ANSWER),
        )
        no_answers = [_row_to_dict(r) for r in cur.fetchall()]
    finally:
        conn.close()

    recent = [
        {
            "interaction_id":   r["interaction_id"],
            "date":             r.get("interaction_date"),
            "call_start_time":  r.get("interaction_call_start_time"),
            "uploaded_at":      r.get("interaction_uploaded_at"),
            "caller_name": (
                ((r.get("user_first_name") or "") + " " +
                 (r.get("user_last_name")  or "")).strip()
                or None
            ),
        }
        for r in no_answers
    ]

    if not row:
        return jsonify({
            "computed":          False,
            "location_id":       location_id,
            "recent_no_answers": recent,
        })

    # Coerce numerics for predictable client-side handling.
    if row.get("li_avg_score") is not None:
        row["li_avg_score"] = float(row["li_avg_score"])
    if row.get("li_last_call_score") is not None:
        row["li_last_call_score"] = float(row["li_last_call_score"])

    row["computed"] = True
    row["location_id"] = location_id
    row["recent_no_answers"] = recent
    return jsonify(row)


# ═══════════════════════════════════════════════════════════════
# DEPARTMENTS
# ═══════════════════════════════════════════════════════════════


@api_bp.route("/departments", methods=["GET"])
@login_required
@role_required("admin", "super_admin")
def list_departments():
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        cur = conn.execute(q("""
            SELECT d.department_id, d.department_name, d.status_id, s.status_name
            FROM departments d
            LEFT JOIN statuses s ON s.status_id = d.status_id
            WHERE d.company_id = ? AND d.department_deleted_at IS NULL
            ORDER BY d.department_name
        """), (company_id,))
        return jsonify(_rows(cur))
    finally:
        conn.close()


@api_bp.route("/departments", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def create_department_route():
    company_id, err = _require_company()
    if err: return err

    body = _body()
    err = _require(body, "department_name")
    if err:
        return _err(err, 400)

    conn = get_conn()
    try:
        department_id = _insert_returning(
            conn,
            sql_pg="""INSERT INTO departments (company_id, department_name, status_id)
                      VALUES (%s, %s, 1) RETURNING department_id""",
            sql_lite="INSERT INTO departments (company_id, department_name, status_id) VALUES (?, ?, 1)",
            params=(company_id, body["department_name"]),
            pk_col="department_id",
        )
        write_audit_log(
            current_user.user_id, ACTION_CREATED, ENTITY_DEPARTMENT, department_id,
            metadata={"company_id": company_id,
                      "department_name": body["department_name"]},
            conn=conn,
        )
        conn.commit()
        return jsonify(_row_to_dict(_get_department(conn, department_id, company_id))), 201
    finally:
        conn.close()


@api_bp.route("/departments/<int:department_id>", methods=["PUT"])
@login_required
@role_required("admin", "super_admin")
def update_department(department_id):
    company_id, err = _require_company()
    if err: return err

    body = _body()
    allowed = {"department_name", "status_id"}
    fields = {k: body[k] for k in allowed if k in body}
    if not fields:
        return _err("No fields to update", 400)

    conn = get_conn()
    try:
        if not _get_department(conn, department_id, company_id):
            return _err("Department not found", 404)

        sets = ", ".join(f"{k} = ?" for k in fields)
        params = list(fields.values()) + [department_id]
        conn.execute(q(f"UPDATE departments SET {sets} WHERE department_id = ?"), params)
        write_audit_log(
            current_user.user_id, ACTION_UPDATED, ENTITY_DEPARTMENT, department_id,
            metadata={"changes": fields}, conn=conn,
        )
        conn.commit()
        return jsonify(_row_to_dict(_get_department(conn, department_id, company_id)))
    finally:
        conn.close()


@api_bp.route("/departments/<int:department_id>", methods=["DELETE"])
@login_required
@role_required("admin", "super_admin")
def delete_department(department_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        if not _get_department(conn, department_id, company_id):
            return _err("Department not found", 404)
        conn.execute(q("""UPDATE departments SET department_deleted_at = CURRENT_TIMESTAMP
                          WHERE department_id = ?"""), (department_id,))
        write_audit_log(
            current_user.user_id, ACTION_DELETED, ENTITY_DEPARTMENT, department_id,
            conn=conn,
        )
        conn.commit()
        return _ok()
    finally:
        conn.close()


@api_bp.route("/departments/<int:department_id>/deletion-impact", methods=["GET"])
@login_required
@role_required("admin", "super_admin")
def department_deletion_impact(department_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        dept = _get_department(conn, department_id, company_id)
        if not dept:
            return _err("Department not found", 404)

        cur = conn.execute(q("""
            SELECT COUNT(*) AS cnt FROM users
            WHERE department_id = ? AND user_deleted_at IS NULL
        """), (department_id,))
        row = cur.fetchone()
        members_count = row["cnt"] if IS_POSTGRES else row[0]

        return jsonify({
            "deletable": True,
            "name": dict(dept).get("department_name"),
            "counts": {"members": members_count},
        })
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
# PHONE ROUTING  (per-location, accessible to all authenticated users)
# ═══════════════════════════════════════════════════════════════


@api_bp.route("/phone_routing", methods=["GET"])
@login_required
def list_phone_routings():
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        cur = conn.execute(q("""
            SELECT phr.phone_routing_id, phr.phone_routing_name, phr.location_id,
                   l.location_name
            FROM phone_routing phr
            JOIN locations l ON l.location_id = phr.location_id
            WHERE l.company_id = ? AND l.location_deleted_at IS NULL
            ORDER BY phr.phone_routing_name
        """), (company_id,))
        return jsonify(_rows(cur))
    finally:
        conn.close()


@api_bp.route("/phone_routing", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def create_phone_routing():
    company_id, err = _require_company()
    if err: return err

    body = _body()
    err = _require(body, "phone_routing_name", "location_id")
    if err:
        return _err(err, 400)

    conn = get_conn()
    try:
        # Verify location belongs to this company
        if not _get_location(conn, body["location_id"], company_id):
            return _err("Location not found in this company", 404)

        phone_routing_id = _insert_returning(
            conn,
            sql_pg="""INSERT INTO phone_routing (location_id, phone_routing_name)
                      VALUES (%s, %s) RETURNING phone_routing_id""",
            sql_lite="INSERT INTO phone_routing (location_id, phone_routing_name) VALUES (?, ?)",
            params=(body["location_id"], body["phone_routing_name"]),
            pk_col="phone_routing_id",
        )
        write_audit_log(
            current_user.user_id, ACTION_CREATED, ENTITY_PHONE_ROUTING, phone_routing_id,
            metadata={"location_id": body["location_id"],
                      "phone_routing_name": body["phone_routing_name"]},
            conn=conn,
        )
        conn.commit()
        return jsonify(_row_to_dict(_get_phone_routing(conn, phone_routing_id, company_id))), 201
    finally:
        conn.close()


@api_bp.route("/phone_routing/<int:phone_routing_id>", methods=["PUT"])
@login_required
@role_required("admin", "super_admin")
def update_phone_routing(phone_routing_id):
    company_id, err = _require_company()
    if err: return err

    body = _body()
    if "phone_routing_name" not in body:
        return _err("No fields to update", 400)

    conn = get_conn()
    try:
        if not _get_phone_routing(conn, phone_routing_id, company_id):
            return _err("Phone routing not found", 404)
        conn.execute(q("UPDATE phone_routing SET phone_routing_name = ? WHERE phone_routing_id = ?"),
                     (body["phone_routing_name"], phone_routing_id))
        write_audit_log(
            current_user.user_id, ACTION_UPDATED, ENTITY_PHONE_ROUTING, phone_routing_id,
            metadata={"phone_routing_name": body["phone_routing_name"]}, conn=conn,
        )
        conn.commit()
        return jsonify(_row_to_dict(_get_phone_routing(conn, phone_routing_id, company_id)))
    finally:
        conn.close()


@api_bp.route("/phone_routing/<int:phone_routing_id>", methods=["DELETE"])
@login_required
@role_required("admin", "super_admin")
def delete_phone_routing(phone_routing_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        if not _get_phone_routing(conn, phone_routing_id, company_id):
            return _err("Phone routing not found", 404)
        try:
            conn.execute(q("DELETE FROM phone_routing WHERE phone_routing_id = ?"), (phone_routing_id,))
            write_audit_log(
                current_user.user_id, ACTION_DELETED, ENTITY_PHONE_ROUTING, phone_routing_id,
                conn=conn,
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            # If projects reference this row, the ON DELETE SET NULL on
            # projects.phone_routing_id should keep this from erroring — but just in case:
            if "foreign key" in str(e).lower():
                return _err("Phone routing is referenced by existing records", 409)
            raise
        return _ok()
    finally:
        conn.close()


@api_bp.route("/phone_routing/<int:phone_routing_id>/deletion-impact", methods=["GET"])
@login_required
@role_required("admin", "super_admin")
def phone_routing_deletion_impact(phone_routing_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        phr = _get_phone_routing(conn, phone_routing_id, company_id)
        if not phr:
            return _err("Phone routing not found", 404)

        cur = conn.execute(q("""
            SELECT COUNT(*) AS cnt FROM projects
            WHERE phone_routing_id = ? AND project_deleted_at IS NULL
        """), (phone_routing_id,))
        row = cur.fetchone()
        projects_count = row["cnt"] if IS_POSTGRES else row[0]

        return jsonify({
            "deletable": True,
            "name": dict(phr).get("phone_routing_name"),
            "counts": {"projects": projects_count},
        })
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
# PROJECTS
# ═══════════════════════════════════════════════════════════════


@api_bp.route("/projects", methods=["GET"])
@login_required
def list_projects():
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        # Location priority: phone_routing → rubric_group. A project without a
        # phone_routing can still be single-location if its rubric group is
        # tied to one, so fall back through the rubric_group location to avoid
        # losing the lock on the grade form.
        cur = conn.execute(q("""
            SELECT p.project_id, p.project_name, p.phone_routing_id, p.rubric_group_id,
                   p.project_start_date, p.project_end_date,
                   p.status_id, s.status_name,
                   p.project_all_locations,
                   phr.phone_routing_name,
                   COALESCE(l_phr.location_id,   l_rubric.location_id)   AS location_id,
                   COALESCE(l_phr.location_name, l_rubric.location_name) AS location_name
            FROM projects p
            LEFT JOIN statuses s ON s.status_id = p.status_id
            LEFT JOIN phone_routing phr ON phr.phone_routing_id = p.phone_routing_id
            LEFT JOIN locations l_phr ON l_phr.location_id = phr.location_id
            LEFT JOIN rubric_groups rg ON rg.rubric_group_id = p.rubric_group_id
            LEFT JOIN locations l_rubric ON l_rubric.location_id = rg.location_id
            WHERE p.company_id = ? AND p.project_deleted_at IS NULL
            ORDER BY p.project_name
        """), (company_id,))
        rows = _rows(cur)
        # For all-locations projects, surface a friendly label and mask the
        # derived location_id so the frontend can't treat it as "this project's"
        # location. The underlying rubric group carries location_id = NULL.
        for r in rows:
            if r.get("project_all_locations"):
                r["location_id"]   = None
                r["location_name"] = "All Locations"
        return jsonify(rows)
    finally:
        conn.close()


@api_bp.route("/projects", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def create_project():
    company_id, err = _require_company()
    if err: return err

    body = _body()
    err = _require(body, "project_name", "rubric_group_id", "project_start_date")
    if err:
        return _err(err, 400)

    conn = get_conn()
    try:
        # Verify rubric_group belongs to this company (not a template)
        if not _get_rubric_group_in_company(conn, body["rubric_group_id"], company_id):
            return _err(
                "rubric_group_id not found in this company (industry templates must be cloned first)",
                404,
            )

        # If phone_routing_id provided, verify it belongs to this company
        phone_routing_id = body.get("phone_routing_id")
        if phone_routing_id is not None and not _get_phone_routing(conn, phone_routing_id, company_id):
            return _err("phone_routing_id not found in this company", 404)

        project_id = _insert_returning(
            conn,
            sql_pg="""INSERT INTO projects
                          (company_id, project_name, phone_routing_id, rubric_group_id,
                           project_start_date, project_end_date, status_id)
                      VALUES (%s, %s, %s, %s, %s, %s, 1)
                      RETURNING project_id""",
            sql_lite="""INSERT INTO projects
                            (company_id, project_name, phone_routing_id, rubric_group_id,
                             project_start_date, project_end_date, status_id)
                        VALUES (?, ?, ?, ?, ?, ?, 1)""",
            params=(
                company_id,
                body["project_name"],
                phone_routing_id,
                body["rubric_group_id"],
                body["project_start_date"],
                body.get("project_end_date"),
            ),
            pk_col="project_id",
        )
        write_audit_log(
            current_user.user_id, ACTION_CREATED, ENTITY_PROJECT, project_id,
            metadata={"project_name": body["project_name"],
                      "rubric_group_id": body["rubric_group_id"],
                      "phone_routing_id": phone_routing_id},
            conn=conn,
        )
        conn.commit()
        return jsonify(_row_to_dict(_get_project(conn, project_id, company_id))), 201
    finally:
        conn.close()


@api_bp.route("/projects/<int:project_id>", methods=["PUT"])
@login_required
@role_required("admin", "super_admin")
def update_project(project_id):
    company_id, err = _require_company()
    if err: return err

    body = _body()
    allowed = {"project_name", "phone_routing_id", "rubric_group_id",
               "project_start_date", "project_end_date", "status_id"}
    fields = {k: body[k] for k in allowed if k in body}
    # Map the frontend 'all_locations' flag onto the backing column name and
    # keep the rubric group's location_id in sync so future joins still work.
    if "all_locations" in body:
        fields["project_all_locations"] = bool(body["all_locations"])
    if not fields:
        return _err("No fields to update", 400)

    conn = get_conn()
    try:
        project = _get_project(conn, project_id, company_id)
        if not project:
            return _err("Project not found", 404)

        if "rubric_group_id" in fields and fields["rubric_group_id"] is not None:
            if not _get_rubric_group_in_company(conn, fields["rubric_group_id"], company_id):
                return _err("rubric_group_id not found in this company", 404)

        if "phone_routing_id" in fields and fields["phone_routing_id"] is not None:
            if not _get_phone_routing(conn, fields["phone_routing_id"], company_id):
                return _err("phone_routing_id not found in this company", 404)

        sets = ", ".join(f"{k} = ?" for k in fields)
        params = list(fields.values()) + [project_id]
        conn.execute(q(f"UPDATE projects SET {sets} WHERE project_id = ?"), params)

        # Flipping all_locations flips the rubric group's location_id too — a
        # shared rubric has location_id = NULL, a single-location rubric has the
        # project's location. Rubric group is owned 1:1 by the project here.
        if "project_all_locations" in fields:
            rg_id = dict(project).get("rubric_group_id")
            if rg_id is not None:
                if fields["project_all_locations"]:
                    conn.execute(q("UPDATE rubric_groups SET location_id = NULL "
                                   "WHERE rubric_group_id = ?"), (rg_id,))
                elif "location_id" in body and body["location_id"]:
                    if not _get_location(conn, body["location_id"], company_id):
                        return _err("location_id not found in this company", 404)
                    conn.execute(q("UPDATE rubric_groups SET location_id = ? "
                                   "WHERE rubric_group_id = ?"),
                                 (body["location_id"], rg_id))
        write_audit_log(
            current_user.user_id, ACTION_UPDATED, ENTITY_PROJECT, project_id,
            metadata={"changes": fields}, conn=conn,
        )
        conn.commit()
        return jsonify(_row_to_dict(_get_project(conn, project_id, company_id)))
    finally:
        conn.close()


@api_bp.route("/projects/<int:project_id>", methods=["DELETE"])
@login_required
@role_required("admin", "super_admin")
def delete_project(project_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        if not _get_project(conn, project_id, company_id):
            return _err("Project not found", 404)
        conn.execute(q("""UPDATE projects SET project_deleted_at = CURRENT_TIMESTAMP
                          WHERE project_id = ?"""), (project_id,))
        write_audit_log(
            current_user.user_id, ACTION_DELETED, ENTITY_PROJECT, project_id,
            conn=conn,
        )
        conn.commit()
        return _ok()
    finally:
        conn.close()


@api_bp.route("/projects/<int:project_id>/deletion-impact", methods=["GET"])
@login_required
@role_required("admin", "super_admin")
def project_deletion_impact(project_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        proj = _get_project(conn, project_id, company_id)
        if not proj:
            return _err("Project not found", 404)

        cur = conn.execute(q("""
            SELECT COUNT(*) AS cnt FROM interactions
            WHERE project_id = ?
              AND interaction_overall_score IS NOT NULL
              AND interaction_deleted_at IS NULL
        """), (project_id,))
        row = cur.fetchone()
        graded_count = row["cnt"] if IS_POSTGRES else row[0]

        return jsonify({
            "deletable": True,
            "name": dict(proj).get("project_name"),
            "counts": {"graded_interactions": graded_count},
        })
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
# GET /api/projects/<id>/locations  — locations with interactions in this project
# ═══════════════════════════════════════════════════════════════
#
# Powers the Hub Dashboard "Export by Location" section. Returns only
# locations that have at least one non-deleted interaction in this project
# (locations with zero calls are uninteresting for export). Counts are
# inclusive of every status — the per-status filtering happens in the
# bulk-export preflight, not here.


@api_bp.route("/projects/<int:project_id>/locations", methods=["GET"])
@login_required
@role_required("admin", "super_admin")
def list_project_locations(project_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        if not _get_project(conn, project_id, company_id):
            return _err("Project not found", 404)

        cur = conn.execute(
            q("""SELECT loc.location_id, loc.location_name,
                        COUNT(i.interaction_id) AS interaction_count
                 FROM interactions i
                 JOIN locations loc ON loc.location_id = i.interaction_location_id
                 WHERE i.project_id = ?
                   AND i.interaction_deleted_at IS NULL
                   AND loc.location_deleted_at IS NULL
                 GROUP BY loc.location_id, loc.location_name
                 ORDER BY loc.location_name"""),
            (project_id,),
        )
        return jsonify(_rows(cur))
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
# CAMPAIGNS  (scoped to a project)
# ═══════════════════════════════════════════════════════════════
#
# Campaigns are a per-project label applied to an interaction at grade time,
# so downstream reports can slice by campaign (e.g. "April 2026 push").
# Interactions reference campaigns via interactions.campaign_id (SET NULL on
# delete). Soft-deleted campaigns remain on the existing interactions for
# historical reporting; the list endpoint filters them out so new grades can't
# pick a tombstoned campaign.


@api_bp.route("/projects/<int:project_id>/campaigns", methods=["GET"])
@login_required
def list_campaigns(project_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        if not _get_project(conn, project_id, company_id):
            return _err("Project not found", 404)
        # Order by last-used (most recent interaction) first, then by creation
        # date for campaigns that have never been used. Matches the default-
        # selection rule on the Grade page (pre-select items[0]).
        cur = conn.execute(q("""
            SELECT c.campaign_id, c.project_id, c.campaign_name,
                   c.campaign_created_at, c.campaign_updated_at,
                   MAX(i.interaction_created_at) AS last_used_at,
                   COUNT(DISTINCT i.interaction_id) AS usage_count
            FROM campaigns c
            LEFT JOIN interactions i
                   ON i.campaign_id = c.campaign_id
                  AND i.interaction_deleted_at IS NULL
            WHERE c.project_id = ? AND c.campaign_deleted_at IS NULL
            GROUP BY c.campaign_id, c.project_id, c.campaign_name,
                     c.campaign_created_at, c.campaign_updated_at
            ORDER BY last_used_at DESC NULLS LAST, c.campaign_created_at DESC
        """) if IS_POSTGRES else q("""
            SELECT c.campaign_id, c.project_id, c.campaign_name,
                   c.campaign_created_at, c.campaign_updated_at,
                   MAX(i.interaction_created_at) AS last_used_at,
                   COUNT(DISTINCT i.interaction_id) AS usage_count
            FROM campaigns c
            LEFT JOIN interactions i
                   ON i.campaign_id = c.campaign_id
                  AND i.interaction_deleted_at IS NULL
            WHERE c.project_id = ? AND c.campaign_deleted_at IS NULL
            GROUP BY c.campaign_id
            ORDER BY last_used_at IS NULL, last_used_at DESC, c.campaign_created_at DESC
        """), (project_id,))
        return jsonify(_rows(cur))
    finally:
        conn.close()


@api_bp.route("/projects/<int:project_id>/campaigns", methods=["POST"])
@login_required
@role_required("admin", "caller", "super_admin")
def create_campaign(project_id):
    company_id, err = _require_company()
    if err: return err

    body = _body()
    name = (body.get("campaign_name") or "").strip()
    if not name:
        return _err("Missing campaign_name", 400)

    conn = get_conn()
    try:
        if not _get_project(conn, project_id, company_id):
            return _err("Project not found", 404)

        # Case-insensitive duplicate guard, scoped to live campaigns in the
        # same project. Tombstoned rows don't block a new creation with the
        # same name.
        dup = conn.execute(
            q("""SELECT campaign_id FROM campaigns
                 WHERE project_id = ? AND campaign_deleted_at IS NULL
                   AND LOWER(campaign_name) = LOWER(?)"""),
            (project_id, name),
        ).fetchone()
        if dup:
            existing_id = dup["campaign_id"] if IS_POSTGRES else dup[0]
            return jsonify(_row_to_dict(
                _get_campaign_in_company(conn, existing_id, company_id)
            )), 200

        campaign_id = _insert_returning(
            conn,
            sql_pg="""INSERT INTO campaigns (project_id, campaign_name)
                      VALUES (%s, %s) RETURNING campaign_id""",
            sql_lite="INSERT INTO campaigns (project_id, campaign_name) VALUES (?, ?)",
            params=(project_id, name),
            pk_col="campaign_id",
        )
        write_audit_log(
            current_user.user_id, ACTION_CREATED, ENTITY_CAMPAIGN, campaign_id,
            metadata={"project_id": project_id, "campaign_name": name},
            conn=conn,
        )
        conn.commit()
        return jsonify(_row_to_dict(
            _get_campaign_in_company(conn, campaign_id, company_id)
        )), 201
    finally:
        conn.close()


@api_bp.route("/campaigns/<int:campaign_id>", methods=["PUT"])
@login_required
@role_required("admin", "super_admin")
def update_campaign(campaign_id):
    """Rename a campaign. Case-insensitive duplicate guard scoped to live
    campaigns in the same project; tombstoned rows don't block."""
    company_id, err = _require_company()
    if err: return err

    body = _body()
    name = (body.get("campaign_name") or "").strip()
    if not name:
        return _err("Missing campaign_name", 400)

    conn = get_conn()
    try:
        camp = _get_campaign_in_company(conn, campaign_id, company_id)
        if not camp:
            return _err("Campaign not found", 404)
        project_id = camp["project_id"]

        dup = conn.execute(
            q("""SELECT campaign_id FROM campaigns
                 WHERE project_id = ? AND campaign_deleted_at IS NULL
                   AND LOWER(campaign_name) = LOWER(?)
                   AND campaign_id <> ?"""),
            (project_id, name, campaign_id),
        ).fetchone()
        if dup:
            return _err("A campaign with that name already exists in this project", 409)

        conn.execute(
            q("UPDATE campaigns SET campaign_name = ? WHERE campaign_id = ?"),
            (name, campaign_id),
        )
        write_audit_log(
            current_user.user_id, ACTION_UPDATED, ENTITY_CAMPAIGN, campaign_id,
            metadata={"campaign_name": name}, conn=conn,
        )
        conn.commit()
        return jsonify(_row_to_dict(
            _get_campaign_in_company(conn, campaign_id, company_id)
        ))
    finally:
        conn.close()


@api_bp.route("/campaigns/<int:campaign_id>", methods=["DELETE"])
@login_required
@role_required("admin", "super_admin")
def delete_campaign(campaign_id):
    """Soft-delete a campaign. Interactions keep their campaign_id so past
    grades stay labeled; the list endpoint hides deleted rows so new grades
    can't pick them. Hard delete would cascade SET NULL on interactions and
    lose historical reporting context."""
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        if not _get_campaign_in_company(conn, campaign_id, company_id):
            return _err("Campaign not found", 404)
        conn.execute(
            q("""UPDATE campaigns SET campaign_deleted_at = CURRENT_TIMESTAMP
                 WHERE campaign_id = ?"""),
            (campaign_id,),
        )
        write_audit_log(
            current_user.user_id, ACTION_DELETED, ENTITY_CAMPAIGN, campaign_id,
            conn=conn,
        )
        conn.commit()
        return _ok()
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
# GET /api/projects/<id>/summary   — project hub page payload
# ═══════════════════════════════════════════════════════════════
#
# Everything the hub page needs in one call: full project fields, joined
# location + phone_routing + rubric group, rollups (call_count, avg_score), the
# last 5 interactions, and top 3 respondents by average score. Tenant-scoped
# via get_effective_company_id(); no_answer calls excluded from avg_score.
STATUS_NO_ANSWER = 44


@api_bp.route("/projects/<int:project_id>/summary", methods=["GET"])
@login_required
def get_project_summary(project_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        if not _get_project(conn, project_id, company_id):
            return _err("Project not found", 404)

        cur = conn.execute(q("""
            SELECT p.project_id, p.project_name, p.phone_routing_id, p.rubric_group_id,
                   p.project_start_date, p.project_end_date, p.status_id,
                   s.status_name,
                   p.project_all_locations,
                   phr.phone_routing_name,
                   COALESCE(l_phr.location_id,    l_rubric.location_id)    AS location_id,
                   COALESCE(l_phr.location_name,  l_rubric.location_name)  AS location_name,
                   COALESCE(l_phr.location_phone, l_rubric.location_phone) AS location_phone,
                   rg.rg_name AS rubric_group_name,
                   rg.rg_grade_target
            FROM projects p
            LEFT JOIN statuses      s  ON s.status_id      = p.status_id
            LEFT JOIN phone_routing phr ON phr.phone_routing_id = p.phone_routing_id
            LEFT JOIN locations     l_phr ON l_phr.location_id = phr.location_id
            LEFT JOIN rubric_groups rg ON rg.rubric_group_id = p.rubric_group_id
            LEFT JOIN locations     l_rubric   ON l_rubric.location_id   = rg.location_id
            WHERE p.project_id = ? AND p.company_id = ?
              AND p.project_deleted_at IS NULL
        """), (project_id, company_id))
        project = _row_to_dict(cur.fetchone())
        if not project:
            return _err("Project not found", 404)
        if project.get("project_all_locations"):
            project["location_id"]    = None
            project["location_name"]  = "All Locations"
            project["location_phone"] = None

        # Monthly stat cards — mirrors /api/dashboard shape so the hub's stat
        # strip lines up with the landing's (Total Calls / Avg Score / Below
        # 5.0 / Unanswered). Scoped to this project. Month-scoped is the
        # canonical framing; Below 5.0 + Unanswered are only meaningful
        # against a bounded window.
        month_start, month_end = _month_bounds()

        cur = conn.execute(q("""
            SELECT
                COUNT(*) AS total_calls,
                AVG(i.interaction_overall_score) AS avg_score,
                COUNT(CASE WHEN i.interaction_overall_score < 5.0 THEN 1 END)
                    AS below_threshold
            FROM interactions i
            WHERE i.project_id = ?
              AND i.interaction_deleted_at IS NULL
              AND i.status_id <> ?
              AND i.interaction_date >= ?
              AND i.interaction_date <  ?
        """), (project_id, STATUS_NO_ANSWER, month_start, month_end))
        scored_row = _row_to_dict(cur.fetchone()) or {}
        stat_avg_raw = scored_row.get("avg_score")
        stat_avg = round(float(stat_avg_raw), 1) if stat_avg_raw is not None else None

        cur = conn.execute(q("""
            SELECT COUNT(*) AS cnt FROM interactions i
            WHERE i.project_id = ?
              AND i.interaction_deleted_at IS NULL
              AND i.status_id = ?
              AND i.interaction_date >= ?
              AND i.interaction_date <  ?
        """), (project_id, STATUS_NO_ANSWER, month_start, month_end))
        noans_row = _row_to_dict(cur.fetchone()) or {}
        no_answer_count = noans_row.get("cnt") or 0

        stat_cards = {
            "total_calls":     scored_row.get("total_calls") or 0,
            "avg_score":       stat_avg,
            "below_threshold": scored_row.get("below_threshold") or 0,
            "no_answer_count": no_answer_count,
        }

        # Recent calls — last 5 for this project. Per-call location comes
        # from the interaction's stamped location_id (source of truth), so
        # all-locations projects show the actual property per row.
        cur = conn.execute(q("""
            SELECT
                i.interaction_id, i.interaction_date, i.interaction_overall_score,
                i.interaction_call_start_time, i.interaction_uploaded_at,
                i.interaction_flags,
                loc.location_name,
                COALESCE(
                    r.respondent_name,
                    NULLIF(TRIM(u.user_first_name || ' ' || u.user_last_name), ''),
                    i.interaction_responder_name
                ) AS respondent_name
            FROM interactions i
            JOIN projects p ON p.project_id = i.project_id
            LEFT JOIN locations   loc ON loc.location_id   = i.interaction_location_id
            LEFT JOIN users       u   ON u.user_id         = i.respondent_user_id
            LEFT JOIN respondents r   ON r.respondent_id   = i.respondent_id
            WHERE i.project_id = ? AND i.interaction_deleted_at IS NULL
            ORDER BY i.interaction_id DESC
            LIMIT 5
        """), (project_id,))
        recent_calls = _rows(cur)

        # Top callers for this project, keyed on respondent_id so same-named
        # respondents at different locations remain distinct cards. Month-
        # scoped ranking matches the global /app dashboard. Each row is
        # enriched with a project-scoped locations roll-up + last_call
        # timestamp, a rolling 30-day trend, and a Performance Reports
        # deep-link by respondent_id (PR is 1:1 with respondent).
        # NULL / empty / 'Name Not Detected' names are excluded.
        rolling_start = date.today() - timedelta(days=30)

        cur = conn.execute(q("""
            SELECT r.respondent_id,
                   TRIM(r.respondent_name) AS respondent_name,
                   r.location_id,
                   AVG(i.interaction_overall_score) AS avg_score,
                   COUNT(*) AS call_count
            FROM interactions i
            JOIN respondents r ON r.respondent_id = i.respondent_id
            WHERE i.project_id = ?
              AND i.interaction_deleted_at IS NULL
              AND i.status_id <> ?
              AND i.interaction_overall_score IS NOT NULL
              AND i.interaction_date >= ?
              AND i.interaction_date <  ?
              AND r.respondent_name IS NOT NULL
              AND TRIM(r.respondent_name) <> ''
              AND TRIM(r.respondent_name) <> 'Name Not Detected'
            GROUP BY r.respondent_id, r.respondent_name, r.location_id
            ORDER BY avg_score DESC
            LIMIT 3
        """), (project_id, STATUS_NO_ANSWER, month_start, month_end))

        top_agents = []
        for row in _rows(cur):
            respondent_id = row["respondent_id"]
            name = row["respondent_name"]
            a = row.get("avg_score")

            # Per-respondent month-scoped detail (project-scoped) → locations + last_call.
            cur2 = conn.execute(q("""
                SELECT
                    l.location_name,
                    i.interaction_date,
                    i.interaction_call_start_time,
                    i.interaction_uploaded_at
                FROM interactions i
                JOIN respondents r ON r.respondent_id = i.respondent_id
                LEFT JOIN locations l ON l.location_id = r.location_id
                WHERE i.project_id = ?
                  AND i.interaction_deleted_at IS NULL
                  AND i.status_id <> ?
                  AND i.interaction_overall_score IS NOT NULL
                  AND i.interaction_date >= ?
                  AND i.interaction_date <  ?
                  AND i.respondent_id = ?
            """), (project_id, STATUS_NO_ANSWER, month_start, month_end, respondent_id))
            month_calls = _rows(cur2)
            locations = _roll_up_locations(month_calls)

            ts_values = []
            for r in month_calls:
                ts = (r.get("interaction_call_start_time")
                      or r.get("interaction_uploaded_at")
                      or r.get("interaction_date"))
                if ts is not None:
                    ts_values.append(ts)
            last_call = max(ts_values) if ts_values else None
            last_call_iso = (last_call.isoformat()
                             if hasattr(last_call, "isoformat")
                             else (str(last_call) if last_call else None))

            # Per-respondent rolling-30-day trend, project-scoped.
            cur3 = conn.execute(q("""
                SELECT
                    i.interaction_date,
                    i.interaction_overall_score
                FROM interactions i
                WHERE i.project_id = ?
                  AND i.interaction_deleted_at IS NULL
                  AND i.status_id <> ?
                  AND i.interaction_overall_score IS NOT NULL
                  AND i.interaction_date >= ?
                  AND i.respondent_id = ?
            """), (project_id, STATUS_NO_ANSWER, rolling_start, respondent_id))
            trend = _trend_for_calls(_rows(cur3))

            # PR is 1:1 with respondent; lookup returns 0 or 1 row.
            cur4 = conn.execute(q("""
                SELECT pr.performance_report_id
                FROM performance_reports pr
                WHERE pr.respondent_id = ?
            """), (respondent_id,))
            report_url = _report_url_for(name, _rows(cur4))

            top_agents.append({
                "respondent_id":   respondent_id,
                "respondent_name": name,
                "avg_score":       round(float(a), 1) if a is not None else None,
                "call_count":      row["call_count"],
                "locations":       locations,
                "trend":           trend,
                "last_call":       last_call_iso,
                "report_url":      report_url,
            })

        return jsonify({
            **project,
            "stat_cards":   stat_cards,
            "recent_calls": recent_calls,
            "top_agents":   top_agents,
        })
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
# POST /api/projects/create-with-rubric   — wizard commit endpoint
# ═══════════════════════════════════════════════════════════════
#
# Single transaction: creates rubric_groups + rubric_items + projects.
# Used by the project-creation wizard in templates/projects.html after
# the reviewer confirms the generated rubric.
_VALID_SCORE_TYPES = {"out_of_10", "yes_no", "yes_no_pending"}
_VALID_GRADE_TARGETS = {"caller", "respondent"}


@api_bp.route("/projects/create-with-rubric", methods=["POST"])
@login_required
@role_required("admin", "super_admin", "manager")
def create_project_with_rubric():
    company_id, err = _require_company()
    if err: return err

    body = _body()
    all_locations = bool(body.get("all_locations"))
    # location_id only required when NOT in all-locations mode; otherwise the
    # project floats across every company location and the rubric group is
    # created with location_id = NULL.
    required = ["project_name", "project_start_date", "rubric_name", "rubric_items"]
    if not all_locations:
        required.insert(1, "location_id")
    err = _require(body, *required)
    if err: return _err(err, 400)

    items = body.get("rubric_items") or []
    if not isinstance(items, list) or not items:
        return _err("rubric_items must be a non-empty array", 400)

    grade_target = (body.get("rubric_grade_target") or "respondent").strip()
    if grade_target not in _VALID_GRADE_TARGETS:
        return _err("Please choose who you're grading: the person who placed the call or the person who answered the call.", 400)

    for idx, it in enumerate(items):
        if not isinstance(it, dict) or not (it.get("name") or "").strip():
            return _err(f"Rubric item #{idx + 1} is missing a name.", 400)
        st = it.get("score_type") or "out_of_10"
        if st not in _VALID_SCORE_TYPES:
            return _err(
                f"Rubric item #{idx + 1} score type must be 1–10 scale, Yes/No, or Yes/No/Pending.",
                400,
            )

    # Import here so the rubric helpers stay co-located with their module.
    from rubrics_routes import _insert_rubric_group, _insert_rubric_item

    conn = get_conn()
    try:
        rubric_location_id = None   # NULL for all-locations rubric groups
        if not all_locations:
            if not _get_location(conn, body["location_id"], company_id):
                return _err("location_id not found in this company", 404)
            rubric_location_id = body["location_id"]

        phone_routing_id = body.get("phone_routing_id")
        if phone_routing_id is not None and not _get_phone_routing(conn, phone_routing_id, company_id):
            return _err("phone_routing_id not found in this company", 404)

        rubric_group_id = _insert_rubric_group(
            conn,
            location_id=rubric_location_id,
            rg_name=(body["rubric_name"] or "").strip() or body["project_name"],
            rg_grade_target=grade_target,
        )
        for idx, it in enumerate(items):
            try:
                weight = float(it.get("weight") or 1.0)
            except (TypeError, ValueError):
                weight = 1.0
            _insert_rubric_item(
                conn,
                rubric_group_id=rubric_group_id,
                ri_name=(it["name"] or "").strip(),
                ri_score_type=it.get("score_type") or "out_of_10",
                ri_weight=weight,
                ri_scoring_guidance=(it.get("scoring_guidance") or "").strip() or None,
                ri_order=idx,
            )

        project_id = _insert_returning(
            conn,
            sql_pg="""INSERT INTO projects
                          (company_id, project_name, phone_routing_id, rubric_group_id,
                           project_start_date, project_end_date, status_id,
                           project_all_locations)
                      VALUES (%s, %s, %s, %s, %s, %s, 1, %s)
                      RETURNING project_id""",
            sql_lite="""INSERT INTO projects
                            (company_id, project_name, phone_routing_id, rubric_group_id,
                             project_start_date, project_end_date, status_id,
                             project_all_locations)
                        VALUES (?, ?, ?, ?, ?, ?, 1, ?)""",
            params=(
                company_id,
                (body["project_name"] or "").strip(),
                phone_routing_id,
                rubric_group_id,
                body["project_start_date"],
                body.get("project_end_date"),
                all_locations,
            ),
            pk_col="project_id",
        )
        write_audit_log(
            current_user.user_id, ACTION_CREATED, ENTITY_PROJECT, project_id,
            metadata={
                "project_name":    body["project_name"],
                "rubric_group_id": rubric_group_id,
                "phone_routing_id": phone_routing_id,
                "all_locations":   all_locations,
                "via":             "wizard",
            },
            conn=conn,
        )
        conn.commit()
        return jsonify({
            "ok":              True,
            "project_id":      project_id,
            "rubric_group_id": rubric_group_id,
        }), 201
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
# TEAM  (users scoped to current company via department)
# ═══════════════════════════════════════════════════════════════


_TEAM_SELECT = """
    SELECT u.user_id, u.user_email, u.user_first_name, u.user_last_name,
           u.status_id, s.status_name, u.user_created_at,
           r.role_name
    FROM users u
    JOIN departments d ON d.department_id = u.department_id
    LEFT JOIN user_roles ur ON ur.user_role_id = u.user_role_id
    LEFT JOIN roles      r  ON r.role_id       = ur.role_id
    LEFT JOIN statuses   s  ON s.status_id     = u.status_id
"""


@api_bp.route("/team", methods=["GET"])
@login_required
@role_required("admin", "super_admin")
def list_team():
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        cur = conn.execute(q(
            _TEAM_SELECT + " WHERE d.company_id = ? AND u.user_deleted_at IS NULL"
                           " ORDER BY u.user_created_at DESC"
        ), (company_id,))
        return jsonify(_rows(cur))
    finally:
        conn.close()


@api_bp.route("/team", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def create_team_member():
    company_id, err = _require_company()
    if err: return err

    body = _body()
    if "password" in body:
        return _err(
            "Password is now server-generated. Remove the password field from your request.",
            400,
        )
    err = _require(body, "user_email", "user_first_name", "user_last_name",
                   "role_name", "department_id")
    if err:
        return _err(err, 400)

    # Privilege-escalation guard: only super_admins can create super_admins
    if body["role_name"] == "super_admin" and not current_user.is_super_admin:
        return _err("Only super admins can assign the super_admin role", 403)

    # Verify department belongs to this company
    conn = get_conn()
    try:
        if not _get_department(conn, body["department_id"], company_id):
            return _err("Department not found in this company", 404)
    finally:
        conn.close()

    temp_password = generate_temp_password()

    try:
        new_user_id = auth.create_user(
            email=body["user_email"],
            password=temp_password,
            role_name=body["role_name"],
            first_name=body["user_first_name"],
            last_name=body["user_last_name"],
            department_id=body["department_id"],
        )
    except ValueError as e:
        msg = str(e)
        if "already exists" in msg.lower():
            return _err(msg, 409)
        return _err(msg, 400)

    # Force password change on first login (spec requirement).
    # Done as a follow-up UPDATE to avoid modifying auth.create_user().
    conn = get_conn()
    try:
        conn.execute(
            q("UPDATE users SET user_must_change_password = ? WHERE user_id = ?"),
            (True, new_user_id),
        )
        write_audit_log(
            current_user.user_id, ACTION_CREATED, ENTITY_USER, new_user_id,
            metadata={"email": body["user_email"],
                      "role_name": body["role_name"],
                      "department_id": body["department_id"]},
            conn=conn,
        )
        conn.commit()
    finally:
        conn.close()

    # Return the new user row in the standard team shape
    conn = get_conn()
    try:
        cur = conn.execute(q(_TEAM_SELECT + " WHERE u.user_id = ?"), (new_user_id,))
        row = cur.fetchone()
        payload = _row_to_dict(row) or {}
        payload["temp_password"] = temp_password
        return jsonify(payload), 201
    finally:
        conn.close()


@api_bp.route("/team/<int:user_id>", methods=["PUT"])
@login_required
@role_required("admin", "super_admin")
def update_team_member(user_id):
    company_id, err = _require_company()
    if err: return err

    body = _body()
    allowed = {"user_first_name", "user_last_name", "user_email", "role_name"}
    fields = {k: body[k] for k in allowed if k in body}
    if not fields:
        return _err("No fields to update", 400)

    # Privilege-escalation guard
    if fields.get("role_name") == "super_admin" and not current_user.is_super_admin:
        return _err("Only super admins can assign the super_admin role", 403)

    conn = get_conn()
    try:
        existing = _get_user_in_company(conn, user_id, company_id)
        if not existing:
            return _err("User not found", 404)

        # Email uniqueness check — only if email is actually changing
        new_email = fields.get("user_email")
        if new_email and new_email.lower() != (existing["user_email"] or "").lower():
            dup = conn.execute(
                q("SELECT 1 FROM users WHERE LOWER(user_email) = LOWER(?) AND user_id <> ?"),
                (new_email, user_id),
            ).fetchone()
            if dup:
                return _err("Email already taken", 409)

        # Handle role change separately — updates user_role_id via user_roles lookup
        role_name = fields.pop("role_name", None)
        if role_name:
            role_row = conn.execute(
                q("SELECT role_id FROM roles WHERE role_name = ?"),
                (role_name,),
            ).fetchone()
            if not role_row:
                return _err(f"Unknown role_name: {role_name}", 400)
            role_id = role_row["role_id"]

            ur_row = conn.execute(
                q("SELECT user_role_id FROM user_roles WHERE role_id = ? LIMIT 1"),
                (role_id,),
            ).fetchone()
            if ur_row:
                user_role_id = ur_row["user_role_id"]
            else:
                user_role_id = _insert_returning(
                    conn,
                    sql_pg="INSERT INTO user_roles (role_id) VALUES (%s) RETURNING user_role_id",
                    sql_lite="INSERT INTO user_roles (role_id) VALUES (?)",
                    params=(role_id,),
                    pk_col="user_role_id",
                )
            fields["user_role_id"] = user_role_id

        if fields:
            sets = ", ".join(f"{k} = ?" for k in fields)
            params = list(fields.values()) + [user_id]
            conn.execute(q(f"UPDATE users SET {sets} WHERE user_id = ?"), params)
            write_audit_log(
                current_user.user_id, ACTION_UPDATED, ENTITY_USER, user_id,
                metadata={"changes": fields, "role_name": role_name}, conn=conn,
            )

        conn.commit()
        cur = conn.execute(q(_TEAM_SELECT + " WHERE u.user_id = ?"), (user_id,))
        return jsonify(_row_to_dict(cur.fetchone()))
    finally:
        conn.close()


@api_bp.route("/team/<int:user_id>/deactivate", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def deactivate_team_member(user_id):
    company_id, err = _require_company()
    if err: return err

    if user_id == current_user.user_id:
        return _err("Cannot deactivate yourself", 400)

    conn = get_conn()
    try:
        if not _get_user_in_company(conn, user_id, company_id):
            return _err("User not found", 404)
        conn.execute(q("UPDATE users SET status_id = 2 WHERE user_id = ?"), (user_id,))
        write_audit_log(
            current_user.user_id, ACTION_DELETED, ENTITY_USER, user_id,
            metadata={"action": "deactivate", "new_status_id": 2}, conn=conn,
        )
        conn.commit()
        return _ok()
    finally:
        conn.close()


@api_bp.route("/team/<int:user_id>/reactivate", methods=["POST"])
@login_required
@role_required("admin", "super_admin")
def reactivate_team_member(user_id):
    company_id, err = _require_company()
    if err: return err

    conn = get_conn()
    try:
        if not _get_user_in_company(conn, user_id, company_id):
            return _err("User not found", 404)
        conn.execute(q("UPDATE users SET status_id = 1 WHERE user_id = ?"), (user_id,))
        write_audit_log(
            current_user.user_id, ACTION_UPDATED, ENTITY_USER, user_id,
            metadata={"action": "reactivate", "new_status_id": 1}, conn=conn,
        )
        conn.commit()
        return _ok()
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════
# INDUSTRIES  (global reference data — read only)
# ═══════════════════════════════════════════════════════════════


@api_bp.route("/industries", methods=["GET"])
@login_required
def list_industries():
    conn = get_conn()
    try:
        cur = conn.execute(q("""
            SELECT industry_id, industry_name
            FROM industries
            WHERE status_id = 1
            ORDER BY industry_name
        """))
        return jsonify(_rows(cur))
    finally:
        conn.close()
